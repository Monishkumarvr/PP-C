#!/usr/bin/env python3
"""
Decision Support System - Main Entry Point
===========================================
Generates decision support analysis from optimizer results.

Usage:
    python run_decision_support.py

Input:
    production_plan_COMPREHENSIVE_test.xlsx

Output:
    production_plan_DECISION_SUPPORT.xlsx
"""

import pandas as pd
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from decision_support.bottleneck_analyzer import BottleneckAnalyzer
from decision_support.atp_calculator import ATPCalculator
from decision_support.order_risk_dashboard import OrderRiskAnalyzer
from decision_support.recommendations_engine import RecommendationsEngine


class DecisionSupportReportGenerator:
    """Generates the decision support Excel report."""

    def __init__(self, comprehensive_output_path: str):
        """
        Args:
            comprehensive_output_path: Path to production_plan_COMPREHENSIVE_test.xlsx
        """
        self.input_path = comprehensive_output_path
        self.colors = {
            'header': '1F4788',
            'subheader': '366092',
            'critical': 'FF4444',
            'high': 'FFA500',
            'medium': 'FFD700',
            'low': '90EE90',
            'white': 'FFFFFF',
            'light_gray': 'F5F5F5'
        }

    def generate_report(self, output_path: str):
        """Generate the complete decision support report."""
        print("\n" + "="*80)
        print("DECISION SUPPORT SYSTEM - REPORT GENERATION")
        print("="*80)
        print(f"\nInput:  {self.input_path}")
        print(f"Output: {output_path}")
        print(f"Time:   {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")

        # Initialize analyzers
        print("📊 Initializing analyzers...\n")

        bottleneck_analyzer = BottleneckAnalyzer(self.input_path)
        atp_calculator = ATPCalculator(self.input_path)
        risk_analyzer = OrderRiskAnalyzer(self.input_path)
        recommendations_engine = RecommendationsEngine(self.input_path)

        # Generate analysis DataFrames
        print("\n📈 Generating analysis reports...\n")

        # 1. Executive Summary
        exec_summary = self._create_executive_summary(
            bottleneck_analyzer, risk_analyzer, recommendations_engine
        )

        # 2. Bottleneck Analysis
        bottleneck_df = bottleneck_analyzer.to_dataframe()
        bottleneck_summary = bottleneck_analyzer.get_summary_dataframe()

        # 3. Order Risk Dashboard
        risk_df = risk_analyzer.to_dataframe()
        risk_by_customer = risk_analyzer.get_risk_by_customer()
        risk_by_week = risk_analyzer.get_risk_by_week()

        # 4. Capacity Availability
        capacity_df = atp_calculator.get_available_capacity()
        capacity_forecast = atp_calculator.get_capacity_forecast()
        capacity_summary = atp_calculator.get_capacity_summary_by_week()

        # 5. ATP Template for New Orders
        # Create input template if it doesn't exist
        atp_calculator.create_input_template()
        # Check orders from input file or use samples
        atp_template = atp_calculator.create_atp_template()

        # 6. Recommendations
        recommendations_df = recommendations_engine.to_dataframe()
        action_plan = recommendations_engine.get_action_plan()

        # Collect all sheets
        sheets = {
            '1_EXECUTIVE_SUMMARY': exec_summary,
            '2_BOTTLENECK_ANALYSIS': bottleneck_df,
            '3_BOTTLENECK_SUMMARY': bottleneck_summary,
            '4_ORDER_RISK': risk_df,
            '5_RISK_BY_CUSTOMER': risk_by_customer,
            '6_RISK_BY_WEEK': risk_by_week,
            '7_CAPACITY_FORECAST': capacity_forecast,
            '8_CAPACITY_BY_RESOURCE': capacity_summary,
            '9_ATP_NEW_ORDERS': atp_template,
            '10_RECOMMENDATIONS': recommendations_df,
            '11_ACTION_PLAN': action_plan
        }

        # Write to Excel
        print(f"\n💾 Writing report to: {output_path}")
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            for sheet_name, df in sheets.items():
                df.to_excel(writer, sheet_name=sheet_name, index=False)
                print(f"  ✓ Created: {sheet_name}")

        # Apply formatting
        print("\n🎨 Applying formatting...")
        self._apply_formatting(output_path, sheets)

        print("\n" + "="*80)
        print("✅ DECISION SUPPORT REPORT GENERATED SUCCESSFULLY!")
        print("="*80)
        print(f"\n📁 Output: {output_path}")
        print(f"📊 Sheets: {len(sheets)} analysis sheets")
        print("\n📋 Report Contents:")
        print("   • Executive Summary - Key metrics and KPIs")
        print("   • Bottleneck Analysis - Capacity constraints")
        print("   • Order Risk Dashboard - At-risk orders by customer/week")
        print("   • Capacity Forecast - Weekly capacity outlook")
        print("   • Capacity by Resource - Available capacity matrix")
        print("   • ATP New Orders - Check feasibility of potential orders")
        print("   • Recommendations - Actionable improvement suggestions")
        print("   • Action Plan - Prioritized immediate actions")
        print("\n💡 To check NEW ORDERS:")
        print("   1. Open sheet '9_ATP_NEW_ORDERS'")
        print("   2. Review sample results or modify Part_Code/Qty/Week")
        print("   3. Re-run this script to update feasibility results\n")

    def _create_executive_summary(self, bottleneck_analyzer, risk_analyzer,
                                   recommendations_engine) -> pd.DataFrame:
        """Create executive summary with key metrics."""
        # Get analysis results
        bottleneck_report = bottleneck_analyzer.analyze()
        risk_summary = risk_analyzer.get_summary()
        recommendations = recommendations_engine.generate()

        # Count recommendations by priority
        critical_recs = sum(1 for r in recommendations if r.priority == 'Critical')
        high_recs = sum(1 for r in recommendations if r.priority == 'High')

        # Build summary table
        records = [
            {'Metric': 'PRODUCTION HEALTH', 'Value': '', 'Status': ''},
            {'Metric': 'Total Bottlenecks Detected', 'Value': bottleneck_report.total_bottlenecks,
             'Status': '🔴 Critical' if bottleneck_report.total_bottlenecks > 10 else
                      '🟡 Warning' if bottleneck_report.total_bottlenecks > 5 else '🟢 OK'},
            {'Metric': 'Weeks with Bottlenecks', 'Value': bottleneck_report.weeks_with_bottlenecks,
             'Status': ''},
            {'Metric': 'Critical Path Resources', 'Value': ', '.join(bottleneck_report.critical_path[:3]),
             'Status': ''},
            {'Metric': '', 'Value': '', 'Status': ''},
            {'Metric': 'ORDER FULFILLMENT', 'Value': '', 'Status': ''},
            {'Metric': 'Total Orders', 'Value': risk_summary.total_orders, 'Status': ''},
            {'Metric': 'Critical Risk Orders', 'Value': risk_summary.critical_orders,
             'Status': '🔴 Action Required' if risk_summary.critical_orders > 0 else '🟢 OK'},
            {'Metric': 'High Risk Orders', 'Value': risk_summary.high_risk_orders,
             'Status': '🟡 Monitor' if risk_summary.high_risk_orders > 0 else '🟢 OK'},
            {'Metric': 'Total At-Risk Quantity', 'Value': risk_summary.total_at_risk_qty, 'Status': ''},
            {'Metric': 'At-Risk Customers', 'Value': len(risk_summary.at_risk_customers), 'Status': ''},
            {'Metric': '', 'Value': '', 'Status': ''},
            {'Metric': 'RECOMMENDATIONS', 'Value': '', 'Status': ''},
            {'Metric': 'Critical Priority Actions', 'Value': critical_recs,
             'Status': '🔴 Immediate' if critical_recs > 0 else '🟢 None'},
            {'Metric': 'High Priority Actions', 'Value': high_recs,
             'Status': '🟡 This Week' if high_recs > 0 else '🟢 None'},
            {'Metric': 'Total Recommendations', 'Value': len(recommendations), 'Status': ''},
        ]

        return pd.DataFrame(records)

    def _apply_formatting(self, output_path: str, sheets: dict):
        """Apply formatting to all sheets."""
        wb = load_workbook(output_path)

        for sheet_name, df in sheets.items():
            ws = wb[sheet_name]

            # Header formatting
            header_fill = PatternFill(start_color=self.colors['header'],
                                     end_color=self.colors['header'],
                                     fill_type='solid')
            header_font = Font(color='FFFFFF', bold=True, size=11)

            # Apply header formatting
            for col in range(1, len(df.columns) + 1):
                cell = ws.cell(row=1, column=col)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')

            # Auto-size columns
            for col in range(1, len(df.columns) + 1):
                max_length = 0
                column_letter = get_column_letter(col)

                for row in range(1, len(df) + 2):
                    cell = ws.cell(row=row, column=col)
                    try:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    except:
                        pass

                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width

            # Apply conditional formatting for risk/priority columns
            self._apply_conditional_formatting(ws, df, sheet_name)

            # Freeze header row
            ws.freeze_panes = 'A2'

        wb.save(output_path)

    def _apply_conditional_formatting(self, ws, df, sheet_name):
        """Apply conditional formatting based on values."""
        # Find columns with risk/priority/severity
        risk_cols = []
        for i, col in enumerate(df.columns, 1):
            if any(keyword in col.lower() for keyword in
                   ['risk', 'priority', 'severity', 'status']):
                risk_cols.append(i)

        # Apply color coding
        for row in range(2, len(df) + 2):
            for col in risk_cols:
                cell = ws.cell(row=row, column=col)
                if cell.value:
                    val = str(cell.value).lower()

                    if any(x in val for x in ['critical', '🔴']):
                        cell.fill = PatternFill(start_color=self.colors['critical'],
                                              end_color=self.colors['critical'],
                                              fill_type='solid')
                        cell.font = Font(color='FFFFFF', bold=True)
                    elif any(x in val for x in ['high', '🟡', 'warning']):
                        cell.fill = PatternFill(start_color=self.colors['high'],
                                              end_color=self.colors['high'],
                                              fill_type='solid')
                        cell.font = Font(bold=True)
                    elif any(x in val for x in ['medium']):
                        cell.fill = PatternFill(start_color=self.colors['medium'],
                                              end_color=self.colors['medium'],
                                              fill_type='solid')
                    elif any(x in val for x in ['low', '🟢', 'ok', 'none']):
                        cell.fill = PatternFill(start_color=self.colors['low'],
                                              end_color=self.colors['low'],
                                              fill_type='solid')


def main():
    """Main entry point for decision support report generation."""
    # Configuration
    comprehensive_output = 'production_plan_COMPREHENSIVE_test.xlsx'
    decision_support_output = 'production_plan_DECISION_SUPPORT.xlsx'

    print("="*80)
    print("DECISION SUPPORT SYSTEM")
    print("Manufacturing Production Planning Analysis")
    print("="*80)

    # Generate report
    generator = DecisionSupportReportGenerator(comprehensive_output)
    generator.generate_report(decision_support_output)


if __name__ == '__main__':
    main()
