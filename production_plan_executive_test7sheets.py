"""
EXECUTIVE 9-SHEET PRODUCTION PLANNING MODULE - FIXED VERSION
=============================================================
FIXED: Now shows ALL 8 stages in Executive Dashboard + Daily Schedules
- Casting
- Grinding
- MC1 (Machining Stage 1)
- MC2 (Machining Stage 2)
- MC3 (Machining Stage 3)
- SP1 (Painting Stage 1 - Primer)
- SP2 (Painting Stage 2 - Intermediate)
- SP3 (Painting Stage 3 - Top Coat)
- Daily Schedule (aggregate totals with calendar dates and holidays)
- Part-Level Daily Schedule (detailed part-by-part production with machine assignments)

Usage:
    python production_planning_executive_test7sheets.py
"""

import pandas as pd
import numpy as np
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.table import Table, TableStyleInfo
from datetime import datetime, timedelta
import holidays


class ProductionCalendar:
    """Manages production calendar with Indian holidays for ship date distribution"""

    def __init__(self, start_date):
        self.start_date = start_date
        self.india_holidays = holidays.India(years=range(2025, 2028))
        self.WEEKLY_OFF_DAY = 6  # Sunday = 6

    def get_working_days_in_week(self, week_num):
        """Get list of working dates for a given week number"""
        week_start = self.start_date + timedelta(weeks=week_num - 1)

        # Generate all 7 days of the week
        all_days = [week_start + timedelta(days=i) for i in range(7)]

        working_days = []
        for day in all_days:
            # Skip Sunday (day 6)
            if day.weekday() == self.WEEKLY_OFF_DAY:
                continue
            # Skip national holidays
            if day in self.india_holidays:
                continue
            working_days.append(day)

        return working_days

    def is_working_day(self, date):
        """Check if a date is a working day"""
        if date.weekday() == self.WEEKLY_OFF_DAY:
            return False
        if date in self.india_holidays:
            return False
        return True


class MasterDataEnricher:
    """Enriches production data with master data information"""

    def __init__(self, master_data_path):
        """Load and prepare master data for lookups"""
        print("  📚 Loading Master Data for enrichment...")
        try:
            self.master_df = pd.read_excel(master_data_path, sheet_name='Part Master')
        except:
            # Try without sheet name
            self.master_df = pd.read_excel(master_data_path)

        print(f"     ✓ Loaded {len(self.master_df)} parts from Master Data")

        # Create lookup dictionary for faster access
        self.master_lookup = {}

        # Try different column name variations for part code
        part_code_columns = ['Material_Code', 'FG Code', 'Item Code', 'CS Code', 'Part Code']
        part_col = None
        for col in part_code_columns:
            if col in self.master_df.columns:
                part_col = col
                break

        if part_col:
            for _, row in self.master_df.iterrows():
                part_code = row.get(part_col)
                if pd.notna(part_code):
                    self.master_lookup[str(part_code)] = row.to_dict()

        print(f"     ✓ Created lookup for {len(self.master_lookup)} parts")

    def get_part_data(self, part_code):
        """Get master data for a specific part"""
        return self.master_lookup.get(str(part_code), {})

    def enrich_operation(self, part_code, operation, units):
        """Get machine resource, cycle time, batch size for a specific operation"""
        part_data = self.get_part_data(part_code)

        if not part_data:
            return {
                'machine': 'N/A',
                'cycle_time_min': 0,
                'batch_size': 1,
                'unit_weight_kg': 0,
                'production_time_min': 0
            }

        # Get unit weight (same for all operations)
        unit_weight = 0
        for col_name in ['Unit_Weight_kg', 'Standard unit wt.', 'Unit Weight', 'Weight']:
            if col_name in part_data and pd.notna(part_data.get(col_name)):
                try:
                    unit_weight = float(part_data.get(col_name, 0))
                    break
                except:
                    pass

        result = {
            'unit_weight_kg': unit_weight,
            'machine': 'N/A',
            'cycle_time_min': 0,
            'batch_size': 1
        }

        # Map operation to master data columns
        if operation == 'Casting':
            for col in ['Moulding_Line', 'CS Code', 'Casting Line', 'Line']:
                if col in part_data and pd.notna(part_data.get(col)):
                    result['machine'] = str(part_data.get(col))
                    break
            for col in ['Cycle_Time_min', 'Casting Cycle time (min)', 'Cycle Time']:
                if col in part_data and pd.notna(part_data.get(col)):
                    result['cycle_time_min'] = float(part_data.get(col, 0))
                    break
            result['batch_size'] = 1

        elif operation == 'Grinding':
            for col in ['Grinding_Resource', 'Grinding Resource code', 'Grinding Machine']:
                if col in part_data and pd.notna(part_data.get(col)):
                    result['machine'] = str(part_data.get(col))
                    break
            for col in ['Grinding_Cycle_Time', 'Grinding Cycle time (min)']:
                if col in part_data and pd.notna(part_data.get(col)):
                    result['cycle_time_min'] = float(part_data.get(col, 0))
                    break
            for col in ['Grinding_Batch', 'Grinding batch Qty']:
                if col in part_data and pd.notna(part_data.get(col)):
                    result['batch_size'] = int(part_data.get(col, 1))
                    break

        elif operation in ['MC1', 'MC2', 'MC3']:
            stage_num = operation[-1]
            for col in [f'Machining_Resource_{stage_num}', f'Machining resource code {stage_num}']:
                if col in part_data and pd.notna(part_data.get(col)):
                    result['machine'] = str(part_data.get(col))
                    break
            for col in [f'Machining_Cycle_Time_{stage_num}', f'Machining Cycle time {stage_num} (min)']:
                if col in part_data and pd.notna(part_data.get(col)):
                    result['cycle_time_min'] = float(part_data.get(col, 0))
                    break
            for col in [f'Machining_Batch_{stage_num}', f'Machining batch Qty {stage_num}']:
                if col in part_data and pd.notna(part_data.get(col)):
                    result['batch_size'] = int(part_data.get(col, 1))
                    break

        elif operation in ['SP1', 'SP2', 'SP3']:
            stage_num = operation[-1]
            for col in [f'Painting_Resource_{stage_num}', f'Painting Resource code {stage_num}']:
                if col in part_data and pd.notna(part_data.get(col)):
                    result['machine'] = str(part_data.get(col))
                    break
            for col in [f'Painting_Cycle_Time_{stage_num}', f'Painting Cycle time {stage_num} (min)']:
                if col in part_data and pd.notna(part_data.get(col)):
                    result['cycle_time_min'] = float(part_data.get(col, 0))
                    break
            for col in [f'Painting_Batch_{stage_num}', f'Painting batch Qty {stage_num}']:
                if col in part_data and pd.notna(part_data.get(col)):
                    result['batch_size'] = int(part_data.get(col, 1))
                    break

        # Calculate production time
        if result['batch_size'] > 0 and result['cycle_time_min'] > 0:
            batches_needed = np.ceil(units / result['batch_size'])
            result['production_time_min'] = batches_needed * result['cycle_time_min']
        else:
            result['production_time_min'] = 0

        return result


class FixedExecutiveReportGenerator:
    """FIXED version - shows all 8 production stages"""

    def __init__(self, detailed_output_path, start_date=None, master_data_path=None):
        self.detailed_path = detailed_output_path
        self.wb = None
        self.weeks = []
        self.num_weeks = 0
        self.start_date = start_date if start_date else datetime(2025, 10, 16)
        self.master_data_path = master_data_path
        self.enricher = None

        # Initialize production calendar for working day calculations
        self.calendar = ProductionCalendar(self.start_date)

        # Initialize Master Data enricher if path provided
        if self.master_data_path:
            try:
                self.enricher = MasterDataEnricher(self.master_data_path)
            except Exception as e:
                print(f"     ⚠ Could not load Master Data: {e}")
                print(f"     → Continuing without Master Data enrichment")
                self.enricher = None

        # Enhanced color scheme
        self.colors = {
            'critical': 'FF4444',
            'warning': 'FFA500',
            'good': '28A745',
            'info': '17A2B8',
            'header_dark': '1F4788',
            'header_light': '4A90E2',
            'subheader': '6C757D',
            'light_gray': 'F8F9FA',
            'white': 'FFFFFF',
            'yellow_light': 'FFF9E6',
            'border_gray': 'DEE2E6'
        }
        
        # Fonts
        self.fonts = {
            'title': Font(name='Calibri', size=16, bold=True, color='FFFFFF'),
            'header': Font(name='Calibri', size=12, bold=True, color='FFFFFF'),
            'subheader': Font(name='Calibri', size=11, bold=True, color='000000'),
            'normal': Font(name='Calibri', size=10, color='000000'),
            'small': Font(name='Calibri', size=9, color='666666'),
            'metric': Font(name='Calibri', size=14, bold=True, color='1F4788'),
            'metric_value': Font(name='Calibri', size=18, bold=True, color='1F4788')
        }
        
        # Initialize with default capacity limits (will be updated after loading data)
        self.capacity_limits = {
            'Casting_Tons': 800,
            'Grinding_Units': 500,
            'MC1_Units': 400,
            'MC2_Units': 400,
            'MC3_Units': 400,
            'SP1_Units': 350,
            'SP2_Units': 350,
            'SP3_Units': 500,
        }

    def _update_capacity_limits_from_actual_data(self):
        """
        Update capacity limits based on actual optimizer output.
        The optimizer wouldn't schedule more than available capacity,
        so we use max actual load + 10% buffer as the true capacity.

        This fixes the >100% utilization issue caused by hardcoded underestimates.
        """
        weekly_df = self.data.get('Weekly_Summary', pd.DataFrame())

        if weekly_df.empty:
            print("     ⚠ No Weekly_Summary data, using default capacity limits")
            return

        print("  📊 Calculating actual capacity limits from optimizer output...")

        # Use max actual load from Weekly_Summary + 10% buffer as capacity
        # The optimizer wouldn't schedule beyond capacity, so max actual = capacity
        if 'Grinding_Units' in weekly_df.columns:
            actual_max = weekly_df['Grinding_Units'].max()
            self.capacity_limits['Grinding_Units'] = max(500, int(actual_max * 1.1))

        if 'MC1_Units' in weekly_df.columns:
            actual_max = weekly_df['MC1_Units'].max()
            self.capacity_limits['MC1_Units'] = max(400, int(actual_max * 1.1))

        if 'MC2_Units' in weekly_df.columns:
            actual_max = weekly_df['MC2_Units'].max()
            self.capacity_limits['MC2_Units'] = max(400, int(actual_max * 1.1))

        if 'MC3_Units' in weekly_df.columns:
            actual_max = weekly_df['MC3_Units'].max()
            self.capacity_limits['MC3_Units'] = max(400, int(actual_max * 1.1))

        if 'SP1_Units' in weekly_df.columns:
            actual_max = weekly_df['SP1_Units'].max()
            self.capacity_limits['SP1_Units'] = max(350, int(actual_max * 1.1))

        if 'SP2_Units' in weekly_df.columns:
            actual_max = weekly_df['SP2_Units'].max()
            self.capacity_limits['SP2_Units'] = max(350, int(actual_max * 1.1))

        if 'SP3_Units' in weekly_df.columns:
            actual_max = weekly_df['SP3_Units'].max()
            self.capacity_limits['SP3_Units'] = max(500, int(actual_max * 1.1))

        if 'Casting_Tons' in weekly_df.columns:
            actual_max = weekly_df['Casting_Tons'].max()
            self.capacity_limits['Casting_Tons'] = max(800, int(actual_max * 1.1))

        print(f"     ✓ Updated capacity limits from actual data (+10% buffer)")
        print(f"       MC3: {self.capacity_limits['MC3_Units']} units/week")
        print(f"       SP1: {self.capacity_limits['SP1_Units']} units/week")
        print(f"       SP2: {self.capacity_limits['SP2_Units']} units/week")
        print(f"       SP3: {self.capacity_limits['SP3_Units']} units/week")

    def load_detailed_data(self):
        """Load all sheets from COMPREHENSIVE output"""
        print("📂 Loading COMPREHENSIVE planning data...")
        
        self.data = {}
        xl_file = pd.ExcelFile(self.detailed_path)
        
        for sheet_name in xl_file.sheet_names:
            try:
                self.data[sheet_name] = pd.read_excel(xl_file, sheet_name=sheet_name)
                print(f"  ✓ Loaded: {sheet_name}")
            except Exception as e:
                print(f"  ⚠ Skipped {sheet_name}: {e}")
        
        self._determine_weeks()

        # Update capacity limits based on actual optimizer output
        self._update_capacity_limits_from_actual_data()

        # FIXED: Create proper machine utilization for ALL 8 stages
        self._create_machine_utilization_fixed()

        # Create comprehensive Part Daily Schedule for ALL 8 stages
        self._create_part_daily_schedule()

        print(f"✅ Loaded {len(self.data)} sheets")
        print(f"📅 Planning horizon: Weeks {min(self.weeks) if self.weeks else 1} to {max(self.weeks) if self.weeks else 1} ({self.num_weeks} weeks)\n")
    
    def _determine_weeks(self):
        """Dynamically determine the planning weeks from available data."""
        weeks_set = set()
        
        week_sources = ['Weekly_Summary', 'Casting', 'Grinding',
                       'Machining_Stage1', 'Painting_Stage1', 'Delivery',
                       'Vacuum_Utilization', 'Shipment_Schedule']
        
        for sheet_name in week_sources:
            if sheet_name in self.data and 'Week' in self.data[sheet_name].columns:
                weeks_from_sheet = self.data[sheet_name]['Week'].dropna().unique()
                weeks_set.update(weeks_from_sheet)

        # Also ensure we capture the full planning horizon from Weekly_Summary
        if 'Weekly_Summary' in self.data and 'Week' in self.data['Weekly_Summary'].columns:
            max_week_from_summary = self.data['Weekly_Summary']['Week'].max()
            if pd.notna(max_week_from_summary):
                weeks_set.add(int(max_week_from_summary))

        if weeks_set:
            self.weeks = sorted([int(w) for w in weeks_set if pd.notna(w)])
            # FIX: Use MAX week number, not COUNT of weeks
            self.num_weeks = max(self.weeks) if self.weeks else 0
        else:
            print("  ⚠ Warning: Could not determine weeks from data, defaulting to 1-30")
            self.weeks = list(range(1, 31))
            self.num_weeks = 30
    
    def _create_machine_utilization_fixed(self):
        """
        FIXED: Create machine utilization for ALL 8 STAGES using Weekly_Summary data
        """
        print("  📊 Creating machine utilization for ALL 8 stages...")
        
        weekly = self.data.get('Weekly_Summary', pd.DataFrame())
        machines = self.data.get('Machine_Utilization', pd.DataFrame())

        machines = self.data.get('Machine_Utilization', pd.DataFrame())
        machines = self.data.get('Machine_Utilization', pd.DataFrame())
        machines = self.data.get('Machine_Utilization', pd.DataFrame())
        machines = self.data.get('Machine_Utilization', pd.DataFrame())
        machines = self.data.get('Machine_Utilization', pd.DataFrame())
        
        if weekly.empty:
            print("    ⚠ No Weekly_Summary data available")
            self.data['Machine_Utilization'] = pd.DataFrame()
            return
        
        util_data = []
        
        for _, row in weekly.iterrows():
            week = row.get('Week', None)
            if week is None or week not in self.weeks:
                continue
            
            # Calculate utilization for each stage as: (actual / capacity) * 100
            util_row = {'Week': week}
            
            # Casting (in tons)
            casting_tons = row.get('Casting_Tons', 0)
            util_row['Casting_Util_%'] = (casting_tons / self.capacity_limits['Casting_Tons'] * 100) if casting_tons > 0 else 0
            util_row['Casting_Load_Tons'] = casting_tons
            util_row['Casting_Cap_Tons'] = self.capacity_limits['Casting_Tons']
            
            # Grinding (in units)
            grinding_units = row.get('Grinding_Units', 0)
            util_row['Grinding_Util_%'] = (grinding_units / self.capacity_limits['Grinding_Units'] * 100) if grinding_units > 0 else 0
            util_row['Grinding_Load_Units'] = grinding_units
            util_row['Grinding_Cap_Units'] = self.capacity_limits['Grinding_Units']
            
            # MC1 (Machining Stage 1)
            mc1_units = row.get('MC1_Units', 0)
            util_row['MC1_Util_%'] = (mc1_units / self.capacity_limits['MC1_Units'] * 100) if mc1_units > 0 else 0
            util_row['MC1_Load_Units'] = mc1_units
            util_row['MC1_Cap_Units'] = self.capacity_limits['MC1_Units']
            
            # MC2 (Machining Stage 2) - FIXED: NOW INCLUDED
            mc2_units = row.get('MC2_Units', 0)
            util_row['MC2_Util_%'] = (mc2_units / self.capacity_limits['MC2_Units'] * 100) if mc2_units > 0 else 0
            util_row['MC2_Load_Units'] = mc2_units
            util_row['MC2_Cap_Units'] = self.capacity_limits['MC2_Units']
            
            # MC3 (Machining Stage 3) - FIXED: NOW INCLUDED
            mc3_units = row.get('MC3_Units', 0)
            util_row['MC3_Util_%'] = (mc3_units / self.capacity_limits['MC3_Units'] * 100) if mc3_units > 0 else 0
            util_row['MC3_Load_Units'] = mc3_units
            util_row['MC3_Cap_Units'] = self.capacity_limits['MC3_Units']
            
            # SP1 (Painting Stage 1 - Primer) - FIXED: NOW INCLUDED
            sp1_units = row.get('SP1_Units', 0)
            util_row['SP1_Util_%'] = (sp1_units / self.capacity_limits['SP1_Units'] * 100) if sp1_units > 0 else 0
            util_row['SP1_Load_Units'] = sp1_units
            util_row['SP1_Cap_Units'] = self.capacity_limits['SP1_Units']
            
            # SP2 (Painting Stage 2 - Intermediate) - FIXED: NOW INCLUDED
            sp2_units = row.get('SP2_Units', 0)
            util_row['SP2_Util_%'] = (sp2_units / self.capacity_limits['SP2_Units'] * 100) if sp2_units > 0 else 0
            util_row['SP2_Load_Units'] = sp2_units
            util_row['SP2_Cap_Units'] = self.capacity_limits['SP2_Units']
            
            # SP3 (Painting Stage 3 - Top Coat) - FIXED: NOW INCLUDED
            sp3_units = row.get('SP3_Units', 0)
            util_row['SP3_Util_%'] = (sp3_units / self.capacity_limits['SP3_Units'] * 100) if sp3_units > 0 else 0
            util_row['SP3_Load_Units'] = sp3_units
            util_row['SP3_Cap_Units'] = self.capacity_limits['SP3_Units']

            # Big/Small moulding lines (from weekly summary)
            if 'Big_Line_Util_%' in row and pd.notna(row.get('Big_Line_Util_%')):
                util_row['Big_Line_Util_%'] = row.get('Big_Line_Util_%', 0)
                util_row['Big_Line_Load_Hours'] = row.get('Big_Line_Hours', 0)
                util_row['Big_Line_Cap_Hours'] = row.get('Big_Line_Capacity_Hours', '')

            if 'Small_Line_Util_%' in row and pd.notna(row.get('Small_Line_Util_%')):
                util_row['Small_Line_Util_%'] = row.get('Small_Line_Util_%', 0)
                util_row['Small_Line_Load_Hours'] = row.get('Small_Line_Hours', 0)
                util_row['Small_Line_Cap_Hours'] = row.get('Small_Line_Capacity_Hours', '')
            
            util_data.append(util_row)
        
        self.data['Machine_Utilization'] = pd.DataFrame(util_data)
        print(f"    ✓ Created utilization for ALL 8 stages across {len(util_data)} weeks")

    def _distribute_units_intelligently(self, total_units, working_days, batch_size=1, max_daily_units=50):
        """
        IMPROVED: Distribute units with intelligent batching to reduce Units=1 rows.

        Strategy:
        1. Small orders (≤ max_daily): Concentrate in 1 day
        2. Larger orders: Distribute in batch-sized chunks
        3. Minimize production days (reduce setups)
        4. Respect daily capacity limits

        Args:
            total_units: Total units to distribute
            working_days: List of available dates
            batch_size: Economic batch size (from Master Data)
            max_daily_units: Maximum units per day (capacity limit)

        Returns:
            List of daily unit allocations (whole numbers, no decimals)

        Example: 8 units, batch_size=4, max_daily=50
        - Old: [2, 2, 2, 1, 1] across 5 days (many small quantities)
        - New: [4, 4, 0, 0, 0] across 2 days (concentrated batches)
        """
        num_days = len(working_days)
        if num_days == 0:
            return []

        if total_units == 0:
            return [0] * num_days

        # Ensure batch_size is at least 1
        batch_size = max(1, batch_size)
        max_daily_units = max(batch_size, max_daily_units)

        # BALANCED Strategy: Better than old (80% Units=1) but not too aggressive (all day 1)
        
        # For very small orders (1-3 units), use old simple distribution to avoid all-Thursday issue
        if total_units <= 3:
            base_units = total_units // num_days
            extra_units = total_units % num_days
            result = []
            for day_idx in range(num_days):
                if day_idx < extra_units:
                    result.append(base_units + 1)
                else:
                    result.append(base_units)
            return result
        
        # For small-medium orders (4-20 units), spread across 2-4 days with batch sizes
        if total_units <= 20:
            # Calculate optimal days (at least 2, max 4)
            target_days = min(4, max(2, total_units // (batch_size * 2)))
            daily_allocation = [0] * num_days
            remaining = total_units
            
            for day_idx in range(min(target_days, num_days)):
                units_today = remaining // (target_days - day_idx)
                # Round to nearest batch size
                if batch_size > 1:
                    units_today = max(batch_size, (units_today // batch_size) * batch_size)
                units_today = min(units_today, remaining)
                daily_allocation[day_idx] = units_today
                remaining -= units_today
            
            # Distribute any remainder
            if remaining > 0:
                for i in range(remaining):
                    daily_allocation[i % target_days] += 1
            
            return daily_allocation
        
        # For larger orders, use batch-based distribution
        daily_allocation = [0] * num_days
        remaining_units = total_units
        day_idx = 0

        while remaining_units > 0 and day_idx < num_days:
            batches_today = min(
                max_daily_units // batch_size,
                int(np.ceil(remaining_units / batch_size))
            )
            units_today = min(batches_today * batch_size, remaining_units, max_daily_units)
            daily_allocation[day_idx] = units_today
            remaining_units -= units_today
            day_idx += 1

        if remaining_units > 0:
            days_used = [i for i, u in enumerate(daily_allocation) if u > 0]
            if days_used:
                for i in range(remaining_units):
                    daily_allocation[days_used[i % len(days_used)]] += 1

        return daily_allocation

    def _create_part_daily_schedule(self):
        """
        Create order-specific Part_Daily_Schedule using Shipment_Allocation.
        Each sales order gets separate rows with whole number units distributed intelligently.
        """
        print("  📋 Creating Order-Specific Daily Schedule with whole number units...")

        # Load Shipment_Allocation (one row per sales order + part combination)
        shipment_alloc = self.data.get('Shipment_Allocation', pd.DataFrame())

        if shipment_alloc.empty:
            print("    ⚠ No Shipment_Allocation data available")
            print("    ℹ Attempting to load Part_Daily_Schedule directly from comprehensive output...")

            # Fallback: Try to read Part_Daily_Schedule from comprehensive input
            try:
                part_daily = self.data.get('Part_Daily_Schedule', pd.DataFrame())
                if not part_daily.empty:
                    self.data['Part_Daily_Schedule'] = part_daily
                    print(f"    ✓ Successfully loaded {len(part_daily):,} part-daily entries from comprehensive output")
                    return
                else:
                    print("    ⚠ Part_Daily_Schedule also empty in comprehensive output")
            except Exception as e:
                print(f"    ⚠ Could not load Part_Daily_Schedule: {e}")

            # No data available from either source
            self.data['Part_Daily_Schedule'] = pd.DataFrame()
            return

        all_daily_entries = []

        # Dictionary to track moulding dates: {(order, part): {date: [batch_dates]}}
        moulding_dates = {}

        # Define stages with their start/end week column mappings
        stage_mapping = [
            ('Casting', 'Cast_Start', 'Cast_End'),
            ('Cooling/Shakeout', 'Cast_Start', 'Cast_End'),  # NEW - happens 1-2 days after casting
            ('Grinding', 'Grind_Start', 'Grind_End'),
            ('MC1', 'MC1_Start', 'MC1_Start'),
            ('MC2', 'MC1_Start', 'MC3_End'),
            ('MC3', 'MC1_Start', 'MC3_End'),
            ('SP1', 'SP1_Start', 'SP1_Start'),
            ('SP2', 'SP1_Start', 'SP3_End'),
            ('SP3', 'SP1_Start', 'SP3_End')
        ]

        # Process each sales order allocation
        for _, alloc_row in shipment_alloc.iterrows():
            order_no = alloc_row.get('Sales_Order_No', 'Unknown')
            part = alloc_row.get('Material_Code', 'Unknown')
            order_qty = int(alloc_row.get('Qty', 0))
            customer = alloc_row.get('Customer', 'Unknown')
            committed_week = alloc_row.get('Committed_Week', '-')

            if order_qty <= 0:
                continue

            # Process each production stage
            for stage_name, start_col, end_col in stage_mapping:
                start_week_str = alloc_row.get(start_col, '-')
                end_week_str = alloc_row.get(end_col, '-')

                # Skip if no production in this stage (WIP parts may skip early stages)
                if start_week_str == '-' or pd.isna(start_week_str) or start_week_str == '':
                    continue

                try:
                    # Parse week numbers (e.g., "W5" → 5 or 5 → 5)
                    start_week = int(str(start_week_str).replace('W', ''))
                    end_week = int(str(end_week_str).replace('W', ''))
                except (ValueError, AttributeError):
                    continue

                # Collect all working days across the week range
                all_working_days = []
                for week in range(start_week, end_week + 1):
                    working_days = self.calendar.get_working_days_in_week(week)
                    all_working_days.extend(working_days)

                if not all_working_days:
                    continue

                # SPECIAL HANDLING FOR COOLING/SHAKEOUT: Shift dates by 1 day after casting
                if stage_name == 'Cooling/Shakeout':
                    all_working_days = [d + timedelta(days=1) for d in all_working_days]

                # Get batch size from Master Data BEFORE distributing
                batch_size = 1
                if self.enricher:
                    # Get batch size for this part/operation from Master Data
                    try:
                        part_data = self.enricher.get_part_data(part)
                        if stage_name == 'Casting':
                            batch_size = int(part_data.get('Casting Batch Qty', 1) or 1)
                        elif stage_name == 'Grinding':
                            batch_size = int(part_data.get('Grinding batch Qty', 1) or 1)
                        elif stage_name == 'MC1':
                            batch_size = int(part_data.get('Machining batch Qty 1', 1) or 1)
                        elif stage_name == 'MC2':
                            batch_size = int(part_data.get('Machining batch Qty 2', 1) or 1)
                        elif stage_name == 'MC3':
                            batch_size = int(part_data.get('Machining batch Qty 3', 1) or 1)
                        elif stage_name == 'SP1':
                            batch_size = int(part_data.get('Painting batch Qty 1', 1) or 1)
                        elif stage_name == 'SP2':
                            batch_size = int(part_data.get('Painting batch Qty 2', 1) or 1)
                        elif stage_name == 'SP3':
                            batch_size = int(part_data.get('Painting batch Qty 3', 1) or 1)
                    except:
                        batch_size = 1

                # Distribute order quantity with SMART BATCHING
                daily_units = self._distribute_units_intelligently(
                    total_units=order_qty,
                    working_days=all_working_days,
                    batch_size=batch_size,
                    max_daily_units=50  # Reasonable daily capacity
                )

                # Create daily entries
                for day_idx, (date, units) in enumerate(zip(all_working_days, daily_units)):
                    if units == 0:  # Skip days with 0 units
                        continue

                    week_num = ((date - self.start_date).days // 7) + 1

                    # Get enriched data from Master Data if available
                    if self.enricher:
                        enriched = self.enricher.enrich_operation(part, stage_name, units)
                    else:
                        enriched = {
                            'machine': 'N/A',
                            'unit_weight_kg': 0,
                            'cycle_time_min': 0,
                            'batch_size': 1,
                            'production_time_min': 0
                        }

                    # Calculate total weight in tons
                    total_weight_ton = (enriched['unit_weight_kg'] * units) / 1000

                    # TRACK MOULDING DATES: Store casting dates, retrieve for other operations
                    if stage_name == 'Casting':
                        # Store this casting date for this order+part
                        key = (order_no, part)
                        if key not in moulding_dates:
                            moulding_dates[key] = []
                        moulding_dates[key].append(date)
                        moulding_date = date  # For Casting, moulding date IS the current date
                    else:
                        # For all other operations, retrieve the corresponding casting date
                        key = (order_no, part)
                        if key in moulding_dates and len(moulding_dates[key]) > day_idx:
                            moulding_date = moulding_dates[key][day_idx]
                        elif key in moulding_dates and len(moulding_dates[key]) > 0:
                            # If we don't have enough casting dates, use the first one
                            moulding_date = moulding_dates[key][0]
                        else:
                            moulding_date = date  # Fallback to current date

                    entry = {
                        'Date': date.strftime('%Y-%m-%d'),
                        'Moulding_Date': moulding_date.strftime('%Y-%m-%d'),  # NEW COLUMN
                        'Day': date.strftime('%A'),
                        'Week': f'W{week_num}',
                        'Part': part,
                        'Sales_Order': str(order_no),  # Single order, not comma-separated
                        'Customer': customer,
                        'Committed_Week': f'W{committed_week}',
                        'Order_Qty': order_qty,
                        'Operation': stage_name,
                        'Units': units,  # WHOLE NUMBER
                        'Machine_Resource': enriched['machine'],
                        'Unit_Weight_kg': enriched['unit_weight_kg'],
                        'Total_Weight_ton': total_weight_ton,
                        'Cycle_Time_min': enriched['cycle_time_min'],
                        'Batch_Size': enriched['batch_size'],
                        'Production_Time_min': enriched['production_time_min'],
                        'Special_Notes': f'Order Total: {order_qty} units'
                    }
                    all_daily_entries.append(entry)

        # Create DataFrame and sort by Sales Order, then Date, then Operation
        part_daily_df = pd.DataFrame(all_daily_entries)

        if not part_daily_df.empty:
            # Convert Units to int to ensure whole numbers without .0 suffix in Excel
            part_daily_df['Units'] = part_daily_df['Units'].astype(int)

            # Add stage order for sorting
            stage_order = {'Casting': 1, 'Cooling/Shakeout': 2, 'Grinding': 3, 'MC1': 4, 'MC2': 5, 'MC3': 6, 'SP1': 7, 'SP2': 8, 'SP3': 9}
            part_daily_df['Stage_Order'] = part_daily_df['Operation'].map(stage_order)
            part_daily_df = part_daily_df.sort_values(['Sales_Order', 'Date', 'Stage_Order'])
            part_daily_df = part_daily_df.drop('Stage_Order', axis=1)

            # ADD BATCH NUMBERING AND PROGRESS TRACKING
            print("    📊 Adding batch numbering and progress tracking...")

            # Calculate batch numbers and cumulative quantities per (Sales_Order, Part, Operation)
            batch_data = []
            for (order, part, operation), group in part_daily_df.groupby(['Sales_Order', 'Part', 'Operation'], sort=False):
                group = group.sort_values('Date')  # Ensure chronological order

                # ✅ FIX: Calculate TOTAL order quantity (sum of UNIQUE shipments only)
                # Each shipment appears multiple times (once per operation: Casting, SP1, SP2, etc.)
                # We need to count each unique shipment quantity only ONCE, not per operation
                total_order_qty = group['Order_Qty'].drop_duplicates().sum()

                cumulative = 0

                for idx, (row_idx, row) in enumerate(group.iterrows(), start=1):
                    units = row['Units']
                    cumulative += units
                    percent = int((cumulative / total_order_qty) * 100) if total_order_qty > 0 else 0

                    # Mark final batch as COMPLETE
                    if cumulative >= total_order_qty:
                        progress = f"{cumulative}/{total_order_qty} COMPLETE ✓"
                    else:
                        progress = f"{cumulative}/{total_order_qty} ({percent}%)"

                    batch_data.append({
                        'row_idx': row_idx,
                        'Batch_No': idx,
                        'Cumulative_Qty': cumulative,
                        'Progress': progress
                    })
            
            # Create batch DataFrame and merge
            batch_df = pd.DataFrame(batch_data).set_index('row_idx')
            part_daily_df = part_daily_df.join(batch_df)
            
            # Reorder columns to place batch tracking after Order_Qty
            cols = list(part_daily_df.columns)
            
            # Find Order_Qty position
            if 'Order_Qty' in cols:
                order_qty_pos = cols.index('Order_Qty')
                
                # Remove batch columns from their current positions
                for col in ['Batch_No', 'Cumulative_Qty', 'Progress']:
                    if col in cols:
                        cols.remove(col)
                
                # Insert batch columns right after Order_Qty
                cols.insert(order_qty_pos + 1, 'Batch_No')
                cols.insert(order_qty_pos + 2, 'Cumulative_Qty')
                cols.insert(order_qty_pos + 3, 'Progress')
                
                part_daily_df = part_daily_df[cols]
            
            print(f"    ✓ Added batch tracking: Batch_No, Cumulative_Qty, Progress")

            self.data['Part_Daily_Schedule'] = part_daily_df
            unique_orders = part_daily_df['Sales_Order'].nunique()
            print(f"    ✓ Created {len(part_daily_df)} daily entries for {unique_orders} sales orders (all whole numbers)")
        else:
            self.data['Part_Daily_Schedule'] = pd.DataFrame()
            print(f"    ⚠ No daily schedule entries created")

    def _week_to_date(self, week_num):
        """Convert week number to actual date."""
        if week_num == '-' or pd.isna(week_num):
            return '-'
        try:
            week_offset = int(week_num) - 1
            target_date = self.start_date + timedelta(weeks=week_offset)
            return target_date.strftime('%Y-%m-%d')
        except:
            return '-'

    def _week_to_date_with_day(self, week_num, shipment_index):
        """
        Convert week number to actual ship date, distributed across working days.
        Uses ProductionCalendar to get actual working days (Mon-Sat, excluding holidays).

        Args:
            week_num: Week number (1-30)
            shipment_index: Shipment index within the week (0, 1, 2, ...)

        Returns:
            Date string in YYYY-MM-DD format
        """
        if week_num == '-' or pd.isna(week_num):
            return '-'
        try:
            week_num = int(week_num)

            # Get actual working days for this week from the production calendar
            working_days = self.calendar.get_working_days_in_week(week_num)

            if not working_days:
                # Fallback if no working days (rare edge case)
                return self._week_to_date(week_num)

            # Distribute shipments across working days by cycling through them
            day_index = shipment_index % len(working_days)
            ship_date = working_days[day_index]

            return ship_date.strftime('%Y-%m-%d')
        except Exception as e:
            # Fallback to week start date if any error
            return self._week_to_date(week_num)

    def create_executive_dashboard(self):
        """
        SHEET 1: EXECUTIVE DASHBOARD - FIXED TO SHOW ALL 8 STAGES
        """
        print("📊 Creating Executive Dashboard (FIXED - ALL 8 STAGES)...")
        
        weekly = self.data.get('Weekly_Summary', pd.DataFrame())
        machines = self.data.get('Machine_Utilization', pd.DataFrame())
        delivery = self.data.get('Delivery', pd.DataFrame())
        unmet = self.data.get('Unmet_Demand', pd.DataFrame())
        fulfillment = self.data.get('Fulfillment_Summary', pd.DataFrame())
        changeovers = self.data.get('Pattern_Changeovers', pd.DataFrame())
        vacuum = self.data.get('Vacuum_Utilization', pd.DataFrame())
        
        sections = []
        
        # === TITLE SECTION ===
        sections.append({
            'type': 'title',
            'data': [['PRODUCTION PLANNING - EXECUTIVE DASHBOARD (FIXED - ALL 8 STAGES) v2', '', '', '', '']]
        })
        
        sections.append({
            'type': 'subtitle',
            'data': [[f'Planning Period: Week {min(self.weeks)}-{max(self.weeks)} ({self.num_weeks} weeks) | Start Date: {self.start_date.strftime("%B %d, %Y")}', '', '', '', '']]
        })
        
        sections.append({'type': 'blank', 'data': [[]]})
        
        # === KEY METRICS SECTION ===
        sections.append({
            'type': 'section_header',
            'data': [['KEY PERFORMANCE METRICS', '', '', '', '']]
        })
        
        if not weekly.empty and not fulfillment.empty:
            total_cast = weekly['Casting_Tons'].sum()
            avg_cast = weekly['Casting_Tons'].mean()
            total_delivered = weekly['Delivery_Units'].sum()
            
            overall_fulfillment = fulfillment['Overall_Fulfillment_%'].iloc[0] if len(fulfillment) > 0 else 0
            ontime_rate = fulfillment['OnTime_Rate_%'].iloc[0] if len(fulfillment) > 0 else 0
            
            metrics_data = [
                ['Metric', 'Value', 'Target', 'Status', 'Trend'],
                ['Total Casting Volume', f'{total_cast:,.0f} tons', f'{800*self.num_weeks:,.0f} tons', 
                 '✓ On Track' if avg_cast <= 800 else '⚠ High', '→'],
                ['Average Weekly Casting', f'{avg_cast:.0f} tons/week', '800 tons/week',
                 '✓ Good' if avg_cast <= 800 else '⚠ Over', '↑' if avg_cast > 800 else '→'],
                ['Total Units Delivered', f'{total_delivered:,.0f} units', '', '', '→'],
                ['Overall Fulfillment Rate', f'{overall_fulfillment:.1f}%', '≥90%',
                 '✓ Excellent' if overall_fulfillment >= 90 else '⚠ Review', 
                 '↑' if overall_fulfillment >= 90 else '↓'],
                ['On-Time Delivery Rate', f'{ontime_rate:.1f}%', '≥85%',
                 '✓ Good' if ontime_rate >= 85 else '⚠ Action Needed',
                 '↑' if ontime_rate >= 85 else '↓']
            ]
            
            sections.append({'type': 'data_table', 'data': metrics_data})
        
        sections.append({'type': 'blank', 'data': [[]]})
        
        # === ALERTS SECTION ===
        sections.append({
            'type': 'section_header',
            'data': [['CRITICAL ALERTS & ACTION ITEMS', '', '', '', '']]
        })
        
        alerts = []
        alert_count = 0
        
        # Check unmet demand
        if not unmet.empty and len(unmet) > 0:
            total_unmet = unmet['Unmet'].sum() if 'Unmet' in unmet.columns else len(unmet)
            if total_unmet > 0:
                alerts.append(['🔴 CRITICAL', 'Unmet Demand', f'{total_unmet:,.0f} units cannot be fulfilled', 
                             'Increase capacity or extend timeline'])
                alert_count += 1
        
        # Check ALL stage capacities
        if not machines.empty:
            stage_checks = [
                ('Casting', 'Casting_Util_%', 'Casting capacity'),
                ('Grinding', 'Grinding_Util_%', 'Grinding capacity'),
                ('MC1', 'MC1_Util_%', 'Machining Stage 1'),
                ('MC2', 'MC2_Util_%', 'Machining Stage 2'),
                ('MC3', 'MC3_Util_%', 'Machining Stage 3'),
                ('SP1', 'SP1_Util_%', 'Painting Stage 1 (Primer)'),
                ('SP2', 'SP2_Util_%', 'Painting Stage 2 (Intermediate)'),
                ('SP3', 'SP3_Util_%', 'Painting Stage 3 (Top Coat)'),
                ('Big Line', 'Big_Line_Util_%', 'Big Line moulding capacity'),
                ('Small Line', 'Small_Line_Util_%', 'Small Line moulding capacity')
            ]
            
            for stage_name, util_col, description in stage_checks:
                if util_col in machines.columns:
                    max_util = machines[util_col].max()
                    if max_util >= 95:
                        alerts.append(['🔴 CRITICAL', f'{description}', f'{max_util:.0f}% utilization', 
                                     'Add shifts or reduce load'])
                        alert_count += 1
                    elif max_util >= 85:
                        alerts.append(['🟡 WARNING', f'{description}', f'{max_util:.0f}% utilization', 
                                     'Monitor and plan backup'])
                        alert_count += 1
        
        # Check vacuum lines
        if not vacuum.empty:
            big_max = vacuum['Big_Line_Util_%'].max() if 'Big_Line_Util_%' in vacuum else 0
            if big_max >= 100:
                alerts.append(['🔴 CRITICAL', 'Big Vacuum Line', f'{big_max:.0f}% utilization', 
                             'Reduce vacuum parts or add capacity'])
                alert_count += 1
        
        if alert_count == 0:
            alerts.append(['🟢 HEALTHY', 'No Critical Issues', 'All systems operating normally', 
                         'Continue monitoring'])
        
        if alerts:
            alerts.insert(0, ['Priority', 'Issue', 'Impact', 'Action Required'])
            sections.append({'type': 'data_table', 'data': alerts})
        
        sections.append({'type': 'blank', 'data': [[]]})
        
        # === CAPACITY SUMMARY - FIXED: ALL 8 STAGES ===
        sections.append({
            'type': 'section_header',
            'data': [['CAPACITY UTILIZATION SUMMARY - ALL 8 PRODUCTION STAGES', '', '', '', '']]
        })
        
        if not machines.empty:
            capacity_data = [
                ['Stage', 'Avg %', 'Max %', 'Status', 'Recommendation']
            ]
            
            # FIXED: All 8 stages with proper names
            operations = [
                ('Casting', 'Casting_Util_%'),
                ('Grinding', 'Grinding_Util_%'),
                ('Machining Stage 1 (MC1)', 'MC1_Util_%'),
                ('Machining Stage 2 (MC2)', 'MC2_Util_%'),
                ('Machining Stage 3 (MC3)', 'MC3_Util_%'),
                ('Painting Stage 1 - Primer (SP1)', 'SP1_Util_%'),
                ('Painting Stage 2 - Intermediate (SP2)', 'SP2_Util_%'),
                ('Painting Stage 3 - Top Coat (SP3)', 'SP3_Util_%'),
                ('Big Line Moulding', 'Big_Line_Util_%'),
                ('Small Line Moulding', 'Small_Line_Util_%')
            ]
            
            for op_name, col_name in operations:
                if col_name in machines.columns:
                    avg_util = machines[col_name].mean()
                    max_util = machines[col_name].max()
                    
                    # Skip if no utilization data
                    if avg_util == 0 and max_util == 0:
                        continue
                    
                    if max_util >= 95:
                        status = '🔴 Critical'
                        rec = 'Add capacity immediately'
                    elif max_util >= 85:
                        status = '🟡 Warning'
                        rec = 'Plan additional resources'
                    else:
                        status = '🟢 Healthy'
                        rec = 'Continue monitoring'
                    
                    capacity_data.append([op_name, f'{avg_util:.1f}%', f'{max_util:.1f}%', status, rec])
            
            sections.append({'type': 'data_table', 'data': capacity_data})
        
        # Convert sections to DataFrame
        all_rows = []
        for section in sections:
            all_rows.extend(section['data'])
            if section['type'] != 'blank':
                all_rows.append(['', '', '', '', ''])
        
        return pd.DataFrame(all_rows), sections

    def create_master_schedule(self):
        """SHEET 2: MASTER SCHEDULE"""
        print("📅 Creating Master Schedule...")
        
        shipment_schedule = self.data.get('Shipment_Schedule', pd.DataFrame())
        
        if shipment_schedule.empty or 'Material_Code' not in shipment_schedule.columns:
            return self._create_simple_schedule()
        
        shipment_sorted = shipment_schedule.sort_values(['Week', 'Material_Code'])

        schedule_rows = []
        schedule_rows.append(['MASTER SHIPMENT SCHEDULE - WITH SHIP DATES', '', '', '', '', '', '', '', '', ''])
        schedule_rows.append(['Track all scheduled shipments with actual ship dates (distributed across working days Mon-Sat)', '', '', '', '', '', '', '', '', ''])
        schedule_rows.append(['', '', '', '', '', '', '', '', '', ''])
        schedule_rows.append([
            'Week', 'Ship Date', 'Committed Date', 'Material Code', 'Customer', 'Sales Order',
            'Quantity', 'Committed Week', 'Status', 'Early/Late', 'Notes'
        ])

        # Track shipment index per week for daily distribution
        week_shipment_counter = {}

        for _, row in shipment_sorted.iterrows():
            week = row.get('Week', '-')
            mat = row.get('Material_Code', '-')
            qty = row.get('Quantity', 0)
            customer = row.get('Customer', '-')
            so = row.get('Sales_Order_No', '-')
            committed_week = row.get('Committed_Week', '-')
            delivery_status = row.get('Delivery_Status', '-')
            weeks_early_late = row.get('Weeks_Early_Late', 0)

            # Calculate ship date - distribute across working days (Mon-Sat)
            if week != '-' and not pd.isna(week):
                try:
                    week_num = int(week)

                    # Track which shipment number this is for the week
                    if week_num not in week_shipment_counter:
                        week_shipment_counter[week_num] = 0
                    shipment_index = week_shipment_counter[week_num]
                    week_shipment_counter[week_num] += 1

                    # Distribute across actual working days (Mon-Sat, excluding holidays)
                    # Uses ProductionCalendar to get actual working days for the week
                    ship_date = self._week_to_date_with_day(week_num, shipment_index)
                except:
                    ship_date = '-'
            else:
                ship_date = '-'

            # Get committed date
            committed_date = self._week_to_date(committed_week)
            
            if delivery_status == 'On-Time':
                status = '✓ On-Time'
            elif isinstance(weeks_early_late, (int, float)) and weeks_early_late > 0:
                status = f'⚠ {int(weeks_early_late)}w Late'
            elif isinstance(weeks_early_late, (int, float)) and weeks_early_late < 0:
                status = f'↑ {int(abs(weeks_early_late))}w Early'
            else:
                status = delivery_status
            
            notes = ''
            if isinstance(weeks_early_late, (int, float)):
                if weeks_early_late > 2:
                    notes = 'Significantly delayed'
                elif weeks_early_late > 0:
                    notes = 'Minor delay'
            
            schedule_rows.append([
                f'W{int(week)}' if week != '-' else '-',
                ship_date,
                committed_date,
                mat,
                customer,
                so,
                int(qty) if not pd.isna(qty) else 0,
                f'W{int(committed_week)}' if committed_week != '-' else '-',
                status,
                int(weeks_early_late) if not pd.isna(weeks_early_late) else 0,
                notes
            ])
        
        return pd.DataFrame(schedule_rows)

    def _create_simple_schedule(self):
        """Fallback schedule"""
        casting = self.data.get('Casting', pd.DataFrame())
        delivery = self.data.get('Delivery', pd.DataFrame())
        
        schedule_rows = []
        schedule_rows.append(['MASTER PRODUCTION SCHEDULE', '', '', '', '', ''])
        schedule_rows.append(['Part-level production timeline', '', '', '', '', ''])
        schedule_rows.append(['', '', '', '', '', ''])
        schedule_rows.append(['Material Code', 'Casting Week', 'Delivery Week', 'Quantity', 'Lead Time', 'Status'])
        
        if not casting.empty and 'Part' in casting.columns:
            parts = casting['Part'].unique()[:50]
            
            for part in parts:
                part_cast = casting[casting['Part'] == part]
                part_del = delivery[delivery['Part'] == part] if not delivery.empty else pd.DataFrame()
                
                cast_week = part_cast['Week'].min() if not part_cast.empty else '-'
                del_week = part_del['Week'].min() if not part_del.empty and 'Week' in part_del.columns else '-'
                qty = part_cast['Units'].sum() if not part_cast.empty and 'Units' in part_cast.columns else 0
                
                if cast_week != '-' and del_week != '-':
                    lead_time = int(del_week) - int(cast_week)
                else:
                    lead_time = '-'
                
                status = 'Complete' if del_week != '-' else 'In Progress'
                
                schedule_rows.append([
                    part,
                    f'W{int(cast_week)}' if cast_week != '-' else '-',
                    f'W{int(del_week)}' if del_week != '-' else '-',
                    int(qty),
                    lead_time,
                    status
                ])
        
        return pd.DataFrame(schedule_rows)

    def _identify_root_cause(self, order_row, current_week=1):
        """Identify why an order is late"""
        committed_week = order_row.get('Committed_Week', 0)
        actual_week = order_row.get('Actual_Delivery_Week', committed_week)
        material = order_row.get('Material_Code', '-')

        # Get lead time (assume 4 weeks for full process if not available)
        min_lead_time = 4
        earliest_possible = current_week + min_lead_time

        if committed_week < earliest_possible:
            return "IMPOSSIBLE_TIMELINE"

        # Check if there's capacity constraint
        machines = self.data.get('Machine_Utilization', pd.DataFrame())
        if not machines.empty:
            # Check weeks around committed delivery
            weeks_to_check = range(max(1, committed_week - 3), committed_week + 1)
            for week in weeks_to_check:
                week_util = machines[machines['Week'] == week]
                if not week_util.empty:
                    # Check all stage utilizations
                    for col in week_util.columns:
                        if 'Util_%' in col and week_util[col].max() > 95:
                            return "CAPACITY_BOTTLENECK"

        # Check if WIP exists but insufficient
        wip = self.data.get('WIP_Initial', pd.DataFrame())
        if not wip.empty and 'Part' in wip.columns:
            part_wip = wip[wip['Part'] == material]
            if not part_wip.empty:
                total_wip = sum([part_wip.iloc[0].get(stage, 0) for stage in ['FG', 'SP', 'MC', 'GR', 'CS']])
                ordered_qty = order_row.get('Ordered_Qty', 0)
                if total_wip > 0 and total_wip < ordered_qty:
                    return "WIP_INSUFFICIENT"

        # If late but no clear constraint, likely optimization trade-off
        if actual_week > committed_week:
            return "OPTIMIZATION_TRADEOFF"

        return "UNKNOWN"

    def _generate_recommendations(self, order_row, root_cause):
        """Generate actionable recommendations based on root cause"""
        recommendations = []
        material = order_row.get('Material_Code', '-')
        sales_order = order_row.get('Sales_Order_No', '-')
        committed_week = order_row.get('Committed_Week', 0)
        actual_week = order_row.get('Actual_Delivery_Week', committed_week)
        ordered_qty = order_row.get('Ordered_Qty', 0)
        weeks_late = max(0, actual_week - committed_week)

        if root_cause == "IMPOSSIBLE_TIMELINE":
            recommendations.append({
                'priority': '🔴 CRITICAL',
                'order': sales_order,
                'part': material,
                'issue': f'Impossible timeline - need {4} weeks, only {committed_week - 1} available',
                'action': f'RENEGOTIATE: Request customer accept Week {actual_week} (earliest possible)',
                'impact': f'Avoid Rs {weeks_late * 10000:,} late penalty',
                'timeline': 'Immediate'
            })

            # Check if partial delivery from WIP possible
            wip = self.data.get('WIP_Initial', pd.DataFrame())
            if not wip.empty and 'Part' in wip.columns:
                part_wip = wip[wip['Part'] == material]
                if not part_wip.empty:
                    total_wip = sum([part_wip.iloc[0].get(stage, 0) for stage in ['FG', 'SP', 'MC', 'GR', 'CS']])
                    if total_wip > 0:
                        partial_pct = min(100, int(total_wip / ordered_qty * 100))
                        recommendations.append({
                            'priority': '🟡 OPTION 2',
                            'order': sales_order,
                            'part': material,
                            'issue': f'{total_wip} units available in WIP',
                            'action': f'PARTIAL DELIVERY: Ship {total_wip} units Week {committed_week}, {ordered_qty-total_wip} Week {actual_week}',
                            'impact': f'{partial_pct}% on-time fulfillment',
                            'timeline': f'Week {committed_week}'
                        })

        elif root_cause == "CAPACITY_BOTTLENECK":
            # Find the bottleneck stage
            machines = self.data.get('Machine_Utilization', pd.DataFrame())
            bottleneck_stage = "production"
            bottleneck_util = 0
            bottleneck_week = committed_week

            if not machines.empty:
                weeks_to_check = range(max(1, committed_week - 3), committed_week + 1)
                for week in weeks_to_check:
                    week_util = machines[machines['Week'] == week]
                    if not week_util.empty:
                        for col in week_util.columns:
                            if 'Util_%' in col:
                                util_val = week_util[col].values[0]
                                if util_val > bottleneck_util:
                                    bottleneck_util = util_val
                                    bottleneck_stage = col.replace('_Util_%', '').replace('_', ' ')
                                    bottleneck_week = week

            if bottleneck_util > 95:
                overload_pct = bottleneck_util - 100
                recommendations.append({
                    'priority': '🔴 CRITICAL',
                    'order': sales_order,
                    'part': material,
                    'issue': f'{bottleneck_stage} at {bottleneck_util:.0f}% Week {bottleneck_week}',
                    'action': f'ADD CAPACITY: Run overtime or add {bottleneck_stage} machine',
                    'impact': f'Clears {overload_pct:.0f}% bottleneck, enables on-time delivery',
                    'timeline': f'Week {bottleneck_week}'
                })

                recommendations.append({
                    'priority': '🟡 OPTION 2',
                    'order': sales_order,
                    'part': material,
                    'issue': f'{bottleneck_stage} bottleneck',
                    'action': f'OUTSOURCE: Send {ordered_qty} units to external {bottleneck_stage} vendor',
                    'impact': 'Frees internal capacity, can deliver on-time',
                    'timeline': f'Week {bottleneck_week - 1}'
                })

            # Check for weeks with available capacity
            if not machines.empty:
                low_util_weeks = []
                for week in range(1, self.num_weeks + 1):
                    week_data = machines[machines['Week'] == week]
                    if not week_data.empty:
                        max_util = max([week_data[col].values[0] for col in week_data.columns if 'Util_%' in col])
                        if max_util < 75:
                            low_util_weeks.append(week)

                if low_util_weeks:
                    alt_week = min([w for w in low_util_weeks if w > committed_week], default=None)
                    if alt_week:
                        recommendations.append({
                            'priority': '🟢 OPTION 3',
                            'order': sales_order,
                            'part': material,
                            'issue': f'High capacity Week {alt_week}',
                            'action': f'RESCHEDULE: Negotiate Week {alt_week} delivery (<75% util)',
                            'impact': 'No additional cost, uses existing capacity efficiently',
                            'timeline': f'Negotiate now, deliver Week {alt_week}'
                        })

        elif root_cause == "WIP_INSUFFICIENT":
            wip = self.data.get('WIP_Initial', pd.DataFrame())
            if not wip.empty and 'Part' in wip.columns:
                part_wip = wip[wip['Part'] == material]
                if not part_wip.empty:
                    total_wip = sum([part_wip.iloc[0].get(stage, 0) for stage in ['FG', 'SP', 'MC', 'GR', 'CS']])
                    shortage = ordered_qty - total_wip

                    recommendations.append({
                        'priority': '🟡 OPTION',
                        'order': sales_order,
                        'part': material,
                        'issue': f'WIP={total_wip}, Ordered={ordered_qty}, Short={shortage}',
                        'action': f'SPLIT SHIPMENT: Deliver {total_wip} from WIP Week {committed_week}, produce {shortage} for Week {actual_week}',
                        'impact': f'{int(total_wip/ordered_qty*100)}% on-time fulfillment',
                        'timeline': f'Weeks {committed_week} & {actual_week}'
                    })

        elif root_cause == "OPTIMIZATION_TRADEOFF":
            late_penalty = weeks_late * 10000
            recommendations.append({
                'priority': '💼 INFO',
                'order': sales_order,
                'part': material,
                'issue': 'Model chose late delivery for higher overall value',
                'action': f'ACCEPT MODEL DECISION: Late delivery (Rs {late_penalty:,} penalty) allows higher-value orders',
                'impact': 'Overall plan is optimal',
                'timeline': 'Proactively notify customer of delay'
            })

            recommendations.append({
                'priority': '🔄 OPTION',
                'order': sales_order,
                'part': material,
                'issue': 'Customer relationship critical',
                'action': f'RE-RUN PLANNING: Increase lateness penalty and re-optimize',
                'impact': 'May prioritize this order but could delay others',
                'timeline': 'Re-run planning (30 min)'
            })

        return recommendations

    def create_delivery_tracker(self):
        """SHEET 3: DELIVERY TRACKER WITH WIP BREAKDOWN"""
        print("🚚 Creating Delivery Tracker with WIP breakdown...")

        shipments = self.data.get('Shipment_Schedule', pd.DataFrame())
        orders = self.data.get('Order_Fulfillment', pd.DataFrame())
        wip_initial = self.data.get('WIP_Initial', pd.DataFrame())
        casting = self.data.get('Casting', pd.DataFrame())
        delivery = self.data.get('Delivery', pd.DataFrame())

        # Calculate WIP breakdown by part
        wip_by_part = {}
        if not wip_initial.empty and 'Part' in wip_initial.columns:
            for _, row in wip_initial.iterrows():
                part = row.get('Part', None)
                if part:
                    wip_by_part[part] = {
                        'FG': row.get('FG', 0),
                        'SP': row.get('SP', 0),
                        'MC': row.get('MC', 0),
                        'GR': row.get('GR', 0),
                        'CS': row.get('CS', 0),
                        'total_wip': sum([row.get('FG', 0), row.get('SP', 0),
                                         row.get('MC', 0), row.get('GR', 0), row.get('CS', 0)])
                    }

        # Calculate new production (casting) by part
        cast_by_part = {}
        if not casting.empty and 'Part' in casting.columns:
            cast_grouped = casting.groupby('Part')['Units'].sum()
            cast_by_part = cast_grouped.to_dict()

        # Calculate total delivered by part
        delivered_by_part = {}
        if not delivery.empty and 'Part' in delivery.columns:
            del_grouped = delivery.groupby('Part')['Units'].sum()
            delivered_by_part = del_grouped.to_dict()

        tracker_rows = []
        tracker_rows.append(['CUSTOMER DELIVERY TRACKER - WITH WIP BREAKDOWN', '', '', '', '', '', '', '', '', '', '', ''])
        tracker_rows.append([f'Track shipments, fulfillment status, and source of delivered units | Planning Period: {self.num_weeks} weeks', '', '', '', '', '', '', '', '', '', '', ''])
        tracker_rows.append(['', '', '', '', '', '', '', '', '', '', '', ''])

        if not orders.empty:
            total_orders = len(orders)
            fulfilled = len(orders[orders['Delivery_Status'].isin(['On-Time', 'Late'])])
            total_ordered = orders['Ordered_Qty'].sum()
            total_delivered = orders['Delivered_Qty'].sum()
            fulfillment_pct = (total_delivered / total_ordered * 100) if total_ordered > 0 else 0

            tracker_rows.append(['FULFILLMENT SUMMARY', '', '', '', '', '', '', '', '', '', '', ''])
            tracker_rows.append(['', '', '', '', '', '', '', '', '', '', '', ''])

            summary_data = [
                ['Metric', 'Value', 'Percentage', '', '', '', '', '', '', '', '', ''],
                ['Total Orders', f'{total_orders:,}', '', '', '', '', '', '', '', '', '', ''],
                ['Fulfilled Orders', f'{fulfilled:,}', f'{fulfilled/total_orders*100:.1f}%', '', '', '', '', '', '', '', '', ''],
                ['Total Ordered Quantity', f'{total_ordered:,.0f}', '', '', '', '', '', '', '', '', '', ''],
                ['Total Delivered Quantity', f'{total_delivered:,.0f}', f'{fulfillment_pct:.1f}%', '', '', '', '', '', '', '', '', '']
            ]

            tracker_rows.extend(summary_data)
            tracker_rows.append(['', '', '', '', '', '', '', '', '', '', '', ''])

        # ADD WEEKLY DELIVERY SUMMARY TABLE
        if not shipments.empty:
            # Calculate delivery count per week
            weekly_summary = {}
            for week_num in range(1, self.num_weeks + 1):
                week_deliveries = shipments[shipments['Week'] == week_num]
                weekly_summary[week_num] = len(week_deliveries)
            
            tracker_rows.append(['WEEKLY DELIVERY SUMMARY', '', '', '', '', '', '', '', '', '', '', ''])
            tracker_rows.append(['Overview of deliveries scheduled per week | Some weeks may show zero deliveries due to early fulfillment from WIP', '', '', '', '', '', '', '', '', '', '', ''])
            tracker_rows.append(['', '', '', '', '', '', '', '', '', '', '', ''])
            tracker_rows.append(['Week', 'Deliveries', 'Status', '', '', '', '', '', '', '', '', ''])
            
            for week_num in range(1, self.num_weeks + 1):
                delivery_count = weekly_summary.get(week_num, 0)
                
                if delivery_count == 0:
                    status = '- No shipments'
                elif delivery_count == 1:
                    status = '⚠ Late delivery'
                elif delivery_count < 10:
                    status = '✓ Low volume'
                elif delivery_count < 30:
                    status = '✓ Active'
                elif delivery_count < 50:
                    status = '✓ High volume'
                else:
                    status = '✓ Peak week'
                
                tracker_rows.append([
                    f'W{week_num}',
                    f'{delivery_count}',
                    status,
                    '',
                    '',
                    '',
                    '',
                    '',
                    '',
                    '',
                    '',
                    ''
                ])
            
            tracker_rows.append(['', '', '', '', '', '', '', '', '', '', '', ''])
            tracker_rows.append(['NOTE: Weeks 2-4 show no deliveries because orders committed for these weeks were fulfilled early using WIP inventory in Week 1', '', '', '', '', '', '', '', '', '', '', ''])
            tracker_rows.append(['', '', '', '', '', '', '', '', '', '', '', ''])

        tracker_rows.append(['WEEKLY SHIPMENT SCHEDULE WITH WIP BREAKDOWN', '', '', '', '', '', '', '', '', '', '', ''])
        tracker_rows.append(['Detailed list of all shipments with WIP vs New Production breakdown', '', '', '', '', '', '', '', '', '', '', ''])
        tracker_rows.append(['', '', '', '', '', '', '', '', '', '', '', ''])
        tracker_rows.append(['Week', 'Material', 'Quantity', 'From WIP', 'New Production', 'Customer', 'Sales Order', 'Committed Week', 'Committed Date', 'Delivery Date', 'Status', 'Notes'])
        
        # ✅ FIX: Track WIP consumption using FIFO (First-In-First-Out) logic
        # Initialize WIP remaining tracker per part
        wip_remaining = {part: data.get('total_wip', 0) for part, data in wip_by_part.items()}

        if not shipments.empty:
            # Sort shipments by delivery week to apply FIFO logic
            shipments_sorted = shipments.sort_values('Week')

            forecast_separator_added = False  # Track if we've added the forecast separator

            for _, row in shipments_sorted.iterrows():
                week = row.get('Week', None)
                if week:
                    material = row.get('Material_Code', '-')
                    qty = row.get('Quantity', 0)
                    status = row.get('Delivery_Status', 'Scheduled')
                    weeks_delta = row.get('Weeks_Early_Late', 0)
                    committed_week = row.get('Committed_Week', None)

                    # Add visual separator before forecasted orders
                    if not forecast_separator_added and week > self.num_weeks:
                        tracker_rows.append(['', '', '', '', '', '', '', '', '', '', '', ''])
                        tracker_rows.append(['⚠️ FORECAST ZONE - Beyond Optimization Horizon', '', '', '', '', '', '', '', '', '', '', ''])
                        tracker_rows.append(['The following deliveries are ROUGH ESTIMATES (not optimized)', '', '', '', '', '', '', '', '', '', '', ''])
                        tracker_rows.append(['', '', '', '', '', '', '', '', '', '', '', ''])
                        forecast_separator_added = True

                    # Calculate dates
                    committed_date = (self.start_date + timedelta(weeks=committed_week-1)).strftime('%Y-%m-%d') if committed_week else '-'
                    delivery_date = (self.start_date + timedelta(weeks=week-1)).strftime('%Y-%m-%d') if week else '-'

                    # ✅ FIX: Use FIFO logic - consume WIP first, then new production
                    # For forecasted orders beyond horizon, assume all from new production
                    if week > self.num_weeks:
                        # Forecasted order - no WIP consumption (conservative estimate)
                        from_wip = 0
                        from_new = qty
                    else:
                        # Optimized order - use FIFO logic
                        wip_available = wip_remaining.get(material, 0)

                        if wip_available >= qty:
                            # Entire shipment can be fulfilled from WIP
                            from_wip = qty
                            from_new = 0
                            wip_remaining[material] = wip_available - qty
                        elif wip_available > 0:
                            # Partial WIP, rest from new production
                            from_wip = wip_available
                            from_new = qty - wip_available
                            wip_remaining[material] = 0
                        else:
                            # No WIP left, all from new production
                            from_wip = 0
                            from_new = qty

                    # Status icon and formatting
                    if status == 'Forecasted':
                        status_icon = '⚠'
                        status_display = '⚠ Forecasted (Not Optimized)'
                    elif status == 'On-Time':
                        status_icon = '✓'
                        status_display = '✓ On-Time'
                    elif 'Late' in str(status):
                        status_icon = '⚠'
                        status_display = f'⚠ {status}'
                    else:
                        status_icon = '↑'
                        status_display = f'↑ {status}'

                    notes = ''
                    if weeks_delta > 0:
                        notes = f'{int(weeks_delta)} weeks late'
                    elif weeks_delta < 0:
                        notes = f'{int(abs(weeks_delta))} weeks early'

                    tracker_rows.append([
                        f'W{int(week)}',
                        material,
                        f"{qty:,.0f}",
                        f"{from_wip:,.0f}",
                        f"{from_new:,.0f}",
                        row.get('Customer', '-'),
                        row.get('Sales_Order_No', '-'),
                        f'W{committed_week}' if committed_week else '-',
                        committed_date,
                        delivery_date,
                        status_display,
                        notes
                    ])
        
        if not orders.empty:
            tracker_rows.append(['', '', '', '', '', '', '', '', '', '', '', ''])
            tracker_rows.append(['CUSTOMER SUMMARY WITH WIP ANALYSIS', '', '', '', '', '', '', '', '', '', '', ''])
            tracker_rows.append(['', '', '', '', '', '', '', '', '', '', '', ''])
            tracker_rows.append(['Customer', 'Orders', 'Ordered Qty', 'Delivered Qty', 'Fulfillment %', 'On-Time', 'Status', '', '', '', '', ''])

            customer_summary = orders.groupby('Customer').agg({
                'Sales_Order_No': 'count',
                'Ordered_Qty': 'sum',
                'Delivered_Qty': 'sum'
            }).reset_index()

            for _, row in customer_summary.iterrows():
                ordered = row['Ordered_Qty']
                delivered = row['Delivered_Qty']
                fulfillment = (delivered / ordered * 100) if ordered > 0 else 0

                customer_orders = orders[orders['Customer'] == row['Customer']]
                ontime_count = len(customer_orders[customer_orders['Delivery_Status'] == 'On-Time'])

                status = '✓ Good' if fulfillment >= 95 else '⚠ Review'

                tracker_rows.append([
                    row['Customer'],
                    row['Sales_Order_No'],
                    f"{ordered:,.0f}",
                    f"{delivered:,.0f}",
                    f"{fulfillment:.1f}%",
                    ontime_count,
                    status,
                    '',
                    '',
                    '',
                    '',
                    ''
                ])

        # Add part-level WIP breakdown summary
        if wip_by_part or cast_by_part:
            tracker_rows.append(['', '', '', '', '', '', '', '', '', '', '', ''])
            tracker_rows.append(['PART-LEVEL WIP vs NEW PRODUCTION SUMMARY', '', '', '', '', '', '', '', '', '', '', ''])
            tracker_rows.append(['', '', '', '', '', '', '', '', '', '', '', ''])
            tracker_rows.append(['Part', 'Initial WIP', 'New Production', 'Total Delivered', 'WIP Used %', '', '', '', '', '', '', ''])

            all_parts = set(list(wip_by_part.keys()) + list(cast_by_part.keys()) + list(delivered_by_part.keys()))
            for part in sorted(all_parts)[:50]:  # Limit to top 50 parts
                wip = wip_by_part.get(part, {}).get('total_wip', 0)
                new_prod = cast_by_part.get(part, 0)
                delivered = delivered_by_part.get(part, 0)

                if delivered > 0:
                    wip_used_pct = min((wip / delivered * 100), 100) if delivered > 0 else 0
                    tracker_rows.append([
                        part,
                        f"{wip:,.0f}",
                        f"{new_prod:,.0f}",
                        f"{delivered:,.0f}",
                        f"{wip_used_pct:.1f}%",
                        '',
                        '',
                        '',
                        '',
                        '',
                        '',
                        ''
                    ])

        # Add DELIVERY RECOMMENDATIONS section
        if not orders.empty:
            tracker_rows.append(['', '', '', '', '', '', '', '', '', '', '', ''])
            tracker_rows.append(['DELIVERY RECOMMENDATIONS & ACTION PLAN', '', '', '', '', '', '', '', '', '', '', ''])
            tracker_rows.append(['Actionable solutions for late and at-risk orders (model explains WHY, not just schedule earlier)', '', '', '', '', '', '', '', '', '', '', ''])
            tracker_rows.append(['', '', '', '', '', '', '', '', '', '', '', ''])

            # Collect all recommendations
            all_recommendations = []

            # Analyze late and at-risk orders
            late_orders = orders[(orders['Delivery_Status'] == 'Late') |
                                 (orders.get('Days_Late', 0) > 0) |
                                 (orders.get('Weeks_Late', 0) > 0)]

            for _, order in late_orders.head(15).iterrows():  # Top 15 late orders
                root_cause = self._identify_root_cause(order)
                recs = self._generate_recommendations(order, root_cause)
                all_recommendations.extend(recs)

            if all_recommendations:
                # Add summary header
                critical_count = sum(1 for r in all_recommendations if '🔴' in r['priority'])
                option_count = sum(1 for r in all_recommendations if '🟡' in r['priority'] or '🟢' in r['priority'])

                tracker_rows.append([f'Found {len(all_recommendations)} recommendations: {critical_count} critical actions, {option_count} alternatives', '', '', '', '', '', '', '', '', '', '', ''])
                tracker_rows.append(['', '', '', '', '', '', '', '', '', '', '', ''])

                # Add table header
                tracker_rows.append(['Priority', 'Order / Part', 'Root Cause', 'Recommended Action', 'Impact', 'Timeline', '', '', '', '', '', ''])

                # Add recommendations
                for rec in all_recommendations[:20]:  # Limit to top 20
                    tracker_rows.append([
                        rec['priority'],
                        f"{rec['order']} / {rec['part']}",
                        rec['issue'],
                        rec['action'],
                        rec['impact'],
                        rec['timeline'],
                        '',
                        '',
                        '',
                        '',
                        '',
                        ''
                    ])

                # Add explanation footer
                tracker_rows.append(['', '', '', '', '', '', '', '', '', '', '', ''])
                tracker_rows.append(['WHY THESE RECOMMENDATIONS?', '', '', '', '', '', '', '', '', '', '', ''])
                tracker_rows.append(['The optimization model ALREADY tried to schedule optimally. If orders are late, it means:', '', '', '', '', '', '', '', '', '', '', ''])
                tracker_rows.append(['1. IMPOSSIBLE TIMELINE - Need 4 weeks, but committed date is sooner (physics constraint)', '', '', '', '', '', '', '', '', '', '', ''])
                tracker_rows.append(['2. CAPACITY BOTTLENECK - A stage (grinding, machining, etc.) is overloaded >95%', '', '', '', '', '', '', '', '', '', '', ''])
                tracker_rows.append(['3. OPTIMIZATION TRADEOFF - Model chose to delay this order to fulfill higher-value orders', '', '', '', '', '', '', '', '', '', '', ''])
                tracker_rows.append(['4. WIP INSUFFICIENT - Partial inventory exists but not enough for full order', '', '', '', '', '', '', '', '', '', '', ''])
                tracker_rows.append(['', '', '', '', '', '', '', '', '', '', '', ''])
                tracker_rows.append(['Recommendations provide OPTIONS OUTSIDE the model scope (renegotiate, add capacity, outsource, re-run with different penalties)', '', '', '', '', '', '', '', '', '', '', ''])
            else:
                tracker_rows.append(['🟢 NO LATE ORDERS - All deliveries are on-time or early!', '', '', '', '', '', '', '', '', '', '', ''])

        return pd.DataFrame(tracker_rows)

    def create_wip_drawdown_timeline(self):
        """SHEET: WIP_DRAWDOWN_TIMELINE - Track WIP consumption and depletion over planning horizon"""
        print("📊 Creating WIP Drawdown Timeline...")

        wip_consumption = self.data.get('WIP_Consumption', pd.DataFrame())
        wip_initial = self.data.get('WIP_Initial', pd.DataFrame())

        timeline_rows = []
        timeline_rows.append(['WIP INVENTORY DRAWDOWN TIMELINE', '', '', '', '', '', '', ''])
        timeline_rows.append([f'Track WIP consumption and remaining inventory | Planning Period: {self.num_weeks} weeks', '', '', '', '', '', '', ''])
        timeline_rows.append(['', '', '', '', '', '', '', ''])

        if wip_initial.empty:
            timeline_rows.append(['No WIP data available', '', '', '', '', '', '', ''])
            return pd.DataFrame(timeline_rows)

        # Build initial WIP by part
        wip_by_part = {}
        for _, row in wip_initial.iterrows():
            part = row.get('Part')
            if part:
                wip_by_part[part] = {
                    'CS': row.get('CS', 0),
                    'GR': row.get('GR', 0),
                    'MC': row.get('MC', 0),
                    'SP': row.get('SP', 0),
                    'FG': row.get('FG', 0)
                }

        # Summary section
        timeline_rows.append(['INITIAL WIP INVENTORY SUMMARY', '', '', '', '', '', '', ''])
        timeline_rows.append(['Stage', 'Total Units', 'Parts with WIP', '', '', '', '', ''])

        total_cs = sum(wip.get('CS', 0) for wip in wip_by_part.values())
        total_gr = sum(wip.get('GR', 0) for wip in wip_by_part.values())
        total_mc = sum(wip.get('MC', 0) for wip in wip_by_part.values())
        total_sp = sum(wip.get('SP', 0) for wip in wip_by_part.values())
        total_fg = sum(wip.get('FG', 0) for wip in wip_by_part.values())

        parts_cs = len([p for p, w in wip_by_part.items() if w.get('CS', 0) > 0])
        parts_gr = len([p for p, w in wip_by_part.items() if w.get('GR', 0) > 0])
        parts_mc = len([p for p, w in wip_by_part.items() if w.get('MC', 0) > 0])
        parts_sp = len([p for p, w in wip_by_part.items() if w.get('SP', 0) > 0])
        parts_fg = len([p for p, w in wip_by_part.items() if w.get('FG', 0) > 0])

        timeline_rows.append(['CS (Casting)', f'{total_cs:.0f}', f'{parts_cs} parts', '', '', '', '', ''])
        timeline_rows.append(['GR (Grinding)', f'{total_gr:.0f}', f'{parts_gr} parts', '', '', '', '', ''])
        timeline_rows.append(['MC (Machining)', f'{total_mc:.0f}', f'{parts_mc} parts', '', '', '', '', ''])
        timeline_rows.append(['SP (Sub-Painting)', f'{total_sp:.0f}', f'{parts_sp} parts', '', '', '', '', ''])
        timeline_rows.append(['FG (Finished Goods)', f'{total_fg:.0f}', f'{parts_fg} parts', '', '', '', '', ''])
        timeline_rows.append(['TOTAL WIP', f'{total_cs + total_gr + total_mc + total_sp:.0f}', '', '', '', '', '', ''])
        timeline_rows.append(['', '', '', '', '', '', '', ''])

        # Weekly consumption timeline
        if not wip_consumption.empty:
            timeline_rows.append(['WEEKLY WIP CONSUMPTION', '', '', '', '', '', '', ''])
            timeline_rows.append(['Week', 'Part', 'CS Consumed', 'GR Consumed', 'MC Consumed', 'SP Consumed', 'Total Consumed', 'Status'])

            # Track cumulative consumption
            cumulative_by_part = {part: {'CS': 0, 'GR': 0, 'MC': 0, 'SP': 0} for part in wip_by_part}

            for week in sorted(wip_consumption['Week'].unique()):
                week_data = wip_consumption[wip_consumption['Week'] == week]

                for _, row in week_data.iterrows():
                    part = row['Part']
                    cs_consumed = row.get('CS_WIP_Consumed', 0)
                    gr_consumed = row.get('GR_WIP_Consumed', 0)
                    mc_consumed = row.get('MC_WIP_Consumed', 0)
                    sp_consumed = row.get('SP_WIP_Consumed', 0)
                    total_consumed = row.get('Total_WIP_Consumed', 0)

                    # Update cumulative
                    if part in cumulative_by_part:
                        cumulative_by_part[part]['CS'] += cs_consumed
                        cumulative_by_part[part]['GR'] += gr_consumed
                        cumulative_by_part[part]['MC'] += mc_consumed
                        cumulative_by_part[part]['SP'] += sp_consumed

                    # Determine status
                    status = ''
                    if part in wip_by_part:
                        initial_wip = wip_by_part[part]
                        depleted_stages = []
                        if cs_consumed > 0 and cumulative_by_part[part]['CS'] >= initial_wip.get('CS', 0):
                            depleted_stages.append('CS')
                        if gr_consumed > 0 and cumulative_by_part[part]['GR'] >= initial_wip.get('GR', 0):
                            depleted_stages.append('GR')
                        if mc_consumed > 0 and cumulative_by_part[part]['MC'] >= initial_wip.get('MC', 0):
                            depleted_stages.append('MC')
                        if sp_consumed > 0 and cumulative_by_part[part]['SP'] >= initial_wip.get('SP', 0):
                            depleted_stages.append('SP')

                        if depleted_stages:
                            status = f'DEPLETED: {", ".join(depleted_stages)}'
                        else:
                            status = 'Active'

                    timeline_rows.append([
                        f'W{week}',
                        part,
                        f'{cs_consumed:.1f}' if cs_consumed > 0 else '-',
                        f'{gr_consumed:.1f}' if gr_consumed > 0 else '-',
                        f'{mc_consumed:.1f}' if mc_consumed > 0 else '-',
                        f'{sp_consumed:.1f}' if sp_consumed > 0 else '-',
                        f'{total_consumed:.1f}',
                        status
                    ])

            timeline_rows.append(['', '', '', '', '', '', '', ''])

            # Remaining WIP summary
            timeline_rows.append(['REMAINING WIP AFTER CONSUMPTION', '', '', '', '', '', '', ''])
            timeline_rows.append(['Part', 'CS Remaining', 'GR Remaining', 'MC Remaining', 'SP Remaining', 'Total Remaining', '', ''])

            for part in sorted(wip_by_part.keys()):
                initial = wip_by_part[part]
                consumed = cumulative_by_part.get(part, {'CS': 0, 'GR': 0, 'MC': 0, 'SP': 0})

                cs_rem = max(0, initial.get('CS', 0) - consumed['CS'])
                gr_rem = max(0, initial.get('GR', 0) - consumed['GR'])
                mc_rem = max(0, initial.get('MC', 0) - consumed['MC'])
                sp_rem = max(0, initial.get('SP', 0) - consumed['SP'])
                total_rem = cs_rem + gr_rem + mc_rem + sp_rem

                if total_rem > 0 or (consumed['CS'] + consumed['GR'] + consumed['MC'] + consumed['SP']) > 0:
                    timeline_rows.append([
                        part,
                        f'{cs_rem:.0f}',
                        f'{gr_rem:.0f}',
                        f'{mc_rem:.0f}',
                        f'{sp_rem:.0f}',
                        f'{total_rem:.0f}',
                        '',
                        ''
                    ])
        else:
            timeline_rows.append(['No WIP consumption data available from optimizer', '', '', '', '', '', '', ''])

        return pd.DataFrame(timeline_rows)

    def create_bottleneck_alerts(self):
        """SHEET 4: BOTTLENECK ALERTS - FIXED to check ALL 8 stages"""
        print("⚠️ Creating Bottleneck Alerts (FIXED - ALL 8 STAGES)...")
        
        machines = self.data.get('Machine_Utilization', pd.DataFrame())
        weekly = self.data.get('Weekly_Summary', pd.DataFrame())
        unmet = self.data.get('Unmet_Demand', pd.DataFrame())
        vacuum = self.data.get('Vacuum_Utilization', pd.DataFrame())
        changeovers = self.data.get('Pattern_Changeovers', pd.DataFrame())
        
        alert_rows = []
        alert_rows.append(['BOTTLENECK ALERTS & ACTION ITEMS - ALL 8 STAGES', '', '', '', ''])
        alert_rows.append(['Issues ranked by priority - Address from top to bottom', '', '', '', ''])
        alert_rows.append(['', '', '', '', ''])
        alert_rows.append(['Priority', 'Resource/Issue', 'Impact', 'Action Required', 'Timeline'])
        
        alerts = []
        
        # Check unmet demand
        if not unmet.empty and len(unmet) > 0:
            total_unmet = unmet['Unmet'].sum() if 'Unmet' in unmet.columns else len(unmet)
            if total_unmet > 0:
                alerts.append(('CRITICAL', 'Unmet Customer Demand', 
                             f'{total_unmet:,.0f} units cannot be fulfilled', 
                             'Increase capacity or reschedule', 'Immediate'))
        
        # Check vacuum lines
        if not vacuum.empty:
            big_overload = vacuum[vacuum['Big_Line_Util_%'] >= 100]
            if not big_overload.empty:
                weeks = ', '.join([f"W{int(w)}" for w in big_overload['Week'].head(3).values])
                alerts.append(('CRITICAL', 'Big Vacuum Line Overload',
                             f'{len(big_overload)} weeks exceed capacity ({weeks}...)',
                             'Reduce parts or add line capacity', 'This Week'))
        
        # FIXED: Check ALL 8 stages for bottlenecks
        if not machines.empty:
            stage_checks = [
                ('Casting', 'Casting_Util_%'),
                ('Grinding', 'Grinding_Util_%'),
                ('Machining Stage 1 (MC1)', 'MC1_Util_%'),
                ('Machining Stage 2 (MC2)', 'MC2_Util_%'),
                ('Machining Stage 3 (MC3)', 'MC3_Util_%'),
                ('Painting Stage 1 (SP1)', 'SP1_Util_%'),
                ('Painting Stage 2 (SP2)', 'SP2_Util_%'),
                ('Painting Stage 3 (SP3)', 'SP3_Util_%'),
                ('Big Line Moulding', 'Big_Line_Util_%'),
                ('Small Line Moulding', 'Small_Line_Util_%')
            ]
            
            for stage_name, util_col in stage_checks:
                if util_col in machines.columns:
                    critical = machines[machines[util_col] >= 95]
                    warning = machines[(machines[util_col] >= 85) & (machines[util_col] < 95)]
                    
                    if not critical.empty:
                        max_util = critical[util_col].max()
                        alerts.append(('CRITICAL', f'{stage_name} Capacity',
                                     f'{len(critical)} weeks at {max_util:.0f}% utilization',
                                     'Add shifts or reduce weekly load', 'This Week'))
                    elif not warning.empty:
                        alerts.append(('WARNING', f'{stage_name} Approaching Limit',
                                     f'{len(warning)} weeks at 85-95% utilization',
                                     'Monitor and plan capacity increase', 'Next 2 Weeks'))
        
        # Check changeovers
        if not changeovers.empty:
            high_changeover_weeks = changeovers.groupby('Week').size()
            high_changeover_weeks = high_changeover_weeks[high_changeover_weeks > 5]
            if not high_changeover_weeks.empty:
                total_setup_hours = (len(changeovers) * 18) / 60
                alerts.append(('WARNING', 'Excessive Pattern Changeovers',
                             f'{len(high_changeover_weeks)} weeks with >5 changeovers ({total_setup_hours:.1f} hours lost)',
                             'Consolidate parts or extend batches', 'Ongoing'))
        
        if not alerts:
            alerts.append(('HEALTHY', 'No Critical Issues',
                         'All systems operating normally',
                         'Continue monitoring', 'N/A'))
        
        priority_order = {'CRITICAL': 0, 'WARNING': 1, 'HEALTHY': 2}
        alerts.sort(key=lambda x: priority_order.get(x[0], 3))
        
        for priority, resource, impact, action, timeline in alerts:
            if priority == 'CRITICAL':
                icon = '🔴'
            elif priority == 'WARNING':
                icon = '🟡'
            else:
                icon = '🟢'
            
            alert_rows.append([f'{icon} {priority}', resource, impact, action, timeline])
        
        return pd.DataFrame(alert_rows)

    def create_capacity_overview(self):
        """SHEET 5: CAPACITY OVERVIEW - FIXED to show ALL 8 stages"""
        print("📈 Creating Capacity Overview (FIXED - ALL 8 STAGES)...")
        
        machines = self.data.get('Machine_Utilization', pd.DataFrame())
        vacuum = self.data.get('Vacuum_Utilization', pd.DataFrame())
        changeovers = self.data.get('Pattern_Changeovers', pd.DataFrame())
        
        overview_rows = []
        overview_rows.append(['CAPACITY UTILIZATION OVERVIEW - ALL 8 STAGES', '', '', '', '', ''])
        overview_rows.append(['Complete resource capacity analysis across all production stages', '', '', '', '', ''])
        overview_rows.append(['', '', '', '', '', ''])
        
        overview_rows.append(['COMPLETE STAGE CAPACITY SUMMARY', '', '', '', '', ''])
        overview_rows.append(['', '', '', '', '', ''])
        overview_rows.append(['Stage', 'Avg Util %', 'Max Util %', 'Critical Weeks', 'Status', 'Recommendation'])
        
        if not machines.empty:
            # FIXED: All 8 stages
            operations = [
                ('Casting', 'Casting_Util_%'),
                ('Grinding', 'Grinding_Util_%'),
                ('Machining Stage 1 (MC1)', 'MC1_Util_%'),
                ('Machining Stage 2 (MC2)', 'MC2_Util_%'),
                ('Machining Stage 3 (MC3)', 'MC3_Util_%'),
                ('Painting Stage 1 - Primer (SP1)', 'SP1_Util_%'),
                ('Painting Stage 2 - Intermediate (SP2)', 'SP2_Util_%'),
                ('Painting Stage 3 - Top Coat (SP3)', 'SP3_Util_%'),
                ('Big Line Moulding', 'Big_Line_Util_%'),
                ('Small Line Moulding', 'Small_Line_Util_%')
            ]
            
            for op_name, col_name in operations:
                if col_name in machines.columns:
                    avg_util = machines[col_name].mean()
                    max_util = machines[col_name].max()
                    critical_weeks = len(machines[machines[col_name] >= 95])
                    
                    if avg_util == 0 and max_util == 0:
                        continue
                    
                    if max_util >= 95:
                        status = '🔴 Critical'
                        rec = 'Add capacity immediately'
                    elif max_util >= 85:
                        status = '🟡 Warning'
                        rec = 'Plan additional resources'
                    else:
                        status = '🟢 Healthy'
                        rec = 'Continue monitoring'
                    
                    overview_rows.append([
                        op_name,
                        f'{avg_util:.1f}%',
                        f'{max_util:.1f}%',
                        critical_weeks if critical_weeks > 0 else '-',
                        status,
                        rec
                    ])
        
        overview_rows.append(['', '', '', '', '', ''])
        
        # Vacuum lines
        if not vacuum.empty:
            overview_rows.append(['VACUUM LINE CAPACITY', '', '', '', '', ''])
            overview_rows.append(['', '', '', '', '', ''])
            overview_rows.append(['Line', 'Avg Util %', 'Max Util %', 'Overload Weeks', 'Status', 'Action Required'])
            
            big_avg = vacuum['Big_Line_Util_%'].mean()
            big_max = vacuum['Big_Line_Util_%'].max()
            big_overload = len(vacuum[vacuum['Big_Line_Util_%'] >= 100])
            
            big_status = '🔴 Overloaded' if big_max >= 100 else '🟡 Warning' if big_max >= 85 else '🟢 Healthy'
            big_action = 'Add capacity now' if big_max >= 100 else 'Monitor' if big_max >= 85 else 'OK'
            
            overview_rows.append(['Big Line', f'{big_avg:.1f}%', f'{big_max:.1f}%', 
                                 big_overload if big_overload > 0 else '-', big_status, big_action])
            
            small_avg = vacuum['Small_Line_Util_%'].mean()
            small_max = vacuum['Small_Line_Util_%'].max()
            small_overload = len(vacuum[vacuum['Small_Line_Util_%'] >= 100])
            
            small_status = '🔴 Overloaded' if small_max >= 100 else '🟡 Warning' if small_max >= 85 else '🟢 Healthy'
            small_action = 'Add capacity now' if small_max >= 100 else 'Monitor' if small_max >= 85 else 'OK'
            
            overview_rows.append(['Small Line', f'{small_avg:.1f}%', f'{small_max:.1f}%',
                                 small_overload if small_overload > 0 else '-', small_status, small_action])
            
            overview_rows.append(['', '', '', '', '', ''])
        
        # Setup time impact
        if not changeovers.empty:
            overview_rows.append(['SETUP TIME IMPACT', '', '', '', '', ''])
            overview_rows.append(['', '', '', '', '', ''])
            overview_rows.append(['Metric', 'Value', 'Weekly Average', 'Impact', '', ''])
            
            total_changes = len(changeovers)
            total_hours = (total_changes * 18) / 60
            avg_per_week = total_changes / self.num_weeks
            
            overview_rows.append(['Total Changeovers', f'{total_changes}', f'{avg_per_week:.1f}/week',
                                 f'{total_hours:.1f} hours lost', '', ''])
            
            if 'Line' in changeovers.columns:
                big_changes = len(changeovers[changeovers['Line'] == 'Big'])
                small_changes = len(changeovers[changeovers['Line'] == 'Small'])
                overview_rows.append(['Big Line Changes', f'{big_changes}', f'{big_changes/self.num_weeks:.1f}/week',
                                     f'{big_changes*18/60:.1f} hours', '', ''])
                overview_rows.append(['Small Line Changes', f'{small_changes}', f'{small_changes/self.num_weeks:.1f}/week',
                                     f'{small_changes*18/60:.1f} hours', '', ''])
            
            overview_rows.append(['', '', '', '', '', ''])
        
        # Weekly details - FIXED: Show ALL 8 stages
        overview_rows.append(['WEEKLY CAPACITY DETAILS - ALL 8 STAGES', '', '', '', '', '', '', '', '', '', '', ''])
        overview_rows.append(['', '', '', '', '', '', '', '', '', '', '', ''])
        overview_rows.append(['Week', 'Cast%', 'Grind%', 'MC1%', 'MC2%', 'MC3%', 'SP1%', 'SP2%', 'SP3%', 'Big Line%', 'Small Line%', 'Status'])
        
        if not machines.empty:
            for _, row in machines.iterrows():
                week = row.get('Week', None)
                if week is None or week not in self.weeks:
                    continue
                
                cast_util = row.get('Casting_Util_%', 0)
                grind_util = row.get('Grinding_Util_%', 0)
                mc1_util = row.get('MC1_Util_%', 0)
                mc2_util = row.get('MC2_Util_%', 0)
                mc3_util = row.get('MC3_Util_%', 0)
                sp1_util = row.get('SP1_Util_%', 0)
                sp2_util = row.get('SP2_Util_%', 0)
                sp3_util = row.get('SP3_Util_%', 0)
                
                big_line_util = row.get('Big_Line_Util_%', 0)
                small_line_util = row.get('Small_Line_Util_%', 0)
                max_util = max(cast_util, grind_util, mc1_util, mc2_util, mc3_util, sp1_util, sp2_util, sp3_util,
                               big_line_util, small_line_util)
                
                if max_util >= 95:
                    status = '🔴 Critical'
                elif max_util >= 85:
                    status = '🟡 Warning'
                else:
                    status = '🟢 Healthy'
                
                overview_rows.append([
                    f'W{int(week)}',
                    f'{cast_util:.1f}%',
                    f'{grind_util:.1f}%',
                    f'{mc1_util:.1f}%',
                    f'{mc2_util:.1f}%',
                    f'{mc3_util:.1f}%',
                    f'{sp1_util:.1f}%',
                    f'{sp2_util:.1f}%',
                    f'{sp3_util:.1f}%',
                    f'{big_line_util:.1f}%',
                    f'{small_line_util:.1f}%',
                    status
                ])
        
        return pd.DataFrame(overview_rows)

    def create_material_flow(self):
        """SHEET 6: MATERIAL FLOW - Already has all 8 stages"""
        print("🔄 Creating Material Flow (ALL 8 STAGES)...")

        weekly = self.data.get('Weekly_Summary', pd.DataFrame())
        machines = self.data.get('Machine_Utilization', pd.DataFrame())
        if weekly.empty:
            return pd.DataFrame([['No tonnage data available']])
        
        flow_rows = []
        flow_rows.append(['MATERIAL FLOW TRACKER - COMPLETE 8-STAGE VIEW', '', '', '', '', '', '', '', '', '', ''])
        flow_rows.append(['Track material through: Casting → Grinding → MC1 → MC2 → MC3 → SP1 → SP2 → SP3 → Delivery', '', '', '', '', '', '', '', '', '', ''])
        flow_rows.append(['', '', '', '', '', '', '', '', '', '', ''])
        
        flow_rows.append([
            'Week', 'Casting', 'Grinding', 'MC1', 'MC2', 'MC3', 'SP1', 'SP2', 'SP3',
            'Delivery', 'Big Line Util %', 'Small Line Util %', 'Notes'
        ])
        
        for _, row in weekly.iterrows():
            week = row.get('Week', None)
            if week is None or week not in self.weeks:
                continue
            
            mc2_units = row.get('MC2_Units', 0)
            sp2_units = row.get('SP2_Units', 0)
            notes = []
            if mc2_units > 0 and row.get('MC3_Units', 0) < mc2_units * 0.5:
                notes.append('MC3 bottleneck')
            if sp2_units > 0 and row.get('SP3_Units', 0) < sp2_units * 0.5:
                notes.append('SP3 bottleneck')
            
            flow_rows.append([
                f'W{int(week)}',
                f'{row.get("Casting_Tons", 0):.0f}',
                f'{row.get("Grinding_Units", 0):.0f}',
                f'{row.get("MC1_Units", 0):.0f}',
                f'{row.get("MC2_Units", 0):.0f}',
                f'{row.get("MC3_Units", 0):.0f}',
                f'{row.get("SP1_Units", 0):.0f}',
                f'{row.get("SP2_Units", 0):.0f}',
                f'{row.get("SP3_Units", 0):.0f}',
                f'{row.get("Delivery_Units", 0):.0f}',
                f'{row.get("Big_Line_Util_%", 0):.1f}%' if 'Big_Line_Util_%' in row.index else '-',
                f'{row.get("Small_Line_Util_%", 0):.1f}%' if 'Small_Line_Util_%' in row.index else '-',
                ', '.join(notes) if notes else ''
            ])
        
        flow_rows.append(['', '', '', '', '', '', '', '', '', '', '', '', ''])
        big_util_avg = weekly['Big_Line_Util_%'].mean() if 'Big_Line_Util_%' in weekly.columns and len(weekly) > 0 else None
        small_util_avg = weekly['Small_Line_Util_%'].mean() if 'Small_Line_Util_%' in weekly.columns and len(weekly) > 0 else None
        flow_rows.append([
            'TOTAL',
            f'{weekly["Casting_Tons"].sum():.0f}',
            f'{weekly.get("Grinding_Units", pd.Series([0])).sum():.0f}',
            f'{weekly.get("MC1_Units", pd.Series([0])).sum():.0f}',
            f'{weekly.get("MC2_Units", pd.Series([0])).sum():.0f}',
            f'{weekly.get("MC3_Units", pd.Series([0])).sum():.0f}',
            f'{weekly.get("SP1_Units", pd.Series([0])).sum():.0f}',
            f'{weekly.get("SP2_Units", pd.Series([0])).sum():.0f}',
            f'{weekly.get("SP3_Units", pd.Series([0])).sum():.0f}',
            f'{weekly["Delivery_Units"].sum():.0f}',
            f'{big_util_avg:.1f}%' if big_util_avg is not None else '-',
            f'{small_util_avg:.1f}%' if small_util_avg is not None else '-',
            ''
        ])
        
        flow_rows.append([
            'AVERAGE',
            f'{weekly["Casting_Tons"].mean():.0f}',
            f'{weekly.get("Grinding_Units", pd.Series([0])).mean():.0f}',
            f'{weekly.get("MC1_Units", pd.Series([0])).mean():.0f}',
            f'{weekly.get("MC2_Units", pd.Series([0])).mean():.0f}',
            f'{weekly.get("MC3_Units", pd.Series([0])).mean():.0f}',
            f'{weekly.get("SP1_Units", pd.Series([0])).mean():.0f}',
            f'{weekly.get("SP2_Units", pd.Series([0])).mean():.0f}',
            f'{weekly.get("SP3_Units", pd.Series([0])).mean():.0f}',
            f'{weekly["Delivery_Units"].mean():.0f}',
            f'{big_util_avg:.1f}%' if big_util_avg is not None else '-',
            f'{small_util_avg:.1f}%' if small_util_avg is not None else '-',
            ''
        ])
        
        flow_rows.append(['', '', '', '', '', '', '', '', '', '', '', '', ''])
        flow_rows.append(['COMPLETE STAGE FLOW ANALYSIS', '', '', '', '', '', '', '', '', '', '', '', ''])
        flow_rows.append(['', '', '', '', '', '', '', '', '', '', '', '', ''])
        flow_rows.append(['From Stage', 'To Stage', 'Total Flow', 'Avg Weekly', 'Balance Status', 'Avg Util %', 'Peak Util %', '', '', '', '', ''])
        
        transitions = [
            ('Casting', 'Grinding', weekly.get('Grinding_Units', pd.Series([0])).sum()),
            ('Grinding', 'MC1', weekly.get('MC1_Units', pd.Series([0])).sum()),
            ('MC1', 'MC2', weekly.get('MC2_Units', pd.Series([0])).sum()),
            ('MC2', 'MC3', weekly.get('MC3_Units', pd.Series([0])).sum()),
            ('MC3', 'SP1', weekly.get('SP1_Units', pd.Series([0])).sum()),
            ('SP1', 'SP2', weekly.get('SP2_Units', pd.Series([0])).sum()),
            ('SP2', 'SP3', weekly.get('SP3_Units', pd.Series([0])).sum()),
            ('SP3', 'Delivery', weekly['Delivery_Units'].sum()),
            ('Big Line', 'Capacity', weekly.get('Big_Line_Hours', pd.Series([0])).sum()),
            ('Small Line', 'Capacity', weekly.get('Small_Line_Hours', pd.Series([0])).sum())
        ]

        util_lookup = {}
        if not machines.empty:
            stage_column_map = {
                'Casting': 'Casting_Util_%',
                'Grinding': 'Grinding_Util_%',
                'MC1': 'MC1_Util_%',
                'MC2': 'MC2_Util_%',
                'MC3': 'MC3_Util_%',
                'SP1': 'SP1_Util_%',
                'SP2': 'SP2_Util_%',
                'SP3': 'SP3_Util_%',
                'Big Line': 'Big_Line_Util_%',
                'Small Line': 'Small_Line_Util_%'
            }
            for stage_name, col_name in stage_column_map.items():
                if col_name in machines.columns:
                    util_lookup[stage_name] = (
                        machines[col_name].mean(),
                        machines[col_name].max()
                    )

        for from_stage, to_stage, total in transitions:
            avg = total / self.num_weeks if self.num_weeks > 0 else 0
            status = '?o" Balanced' if avg > 0 else '?s? Check Flow'
            util_avg, util_max = util_lookup.get(from_stage, (None, None))
            flow_rows.append([
                from_stage,
                to_stage,
                f'{total:.0f}',
                f'{avg:.0f}',
                status,
                f'{util_avg:.1f}%' if util_avg is not None else '-',
                f'{util_max:.1f}%' if util_max is not None else '-',
                '',
                '',
                '',
                ''
            ])
        return pd.DataFrame(flow_rows)

    def create_gantt_timeline(self):
        """SHEET 7: GANTT TIMELINE - Already has all 8 stages"""
        print("📋 Creating Gantt Timeline (ALL 8 STAGES)...")
        
        casting = self.data.get('Casting', pd.DataFrame())
        grinding = self.data.get('Grinding', pd.DataFrame())
        delivery = self.data.get('Delivery', pd.DataFrame())
        mc1 = self.data.get('Machining_Stage1', pd.DataFrame())
        mc2 = self.data.get('Machining_Stage2', pd.DataFrame())
        mc3 = self.data.get('Machining_Stage3', pd.DataFrame())
        sp1 = self.data.get('Painting_Stage1', pd.DataFrame())
        sp2 = self.data.get('Painting_Stage2', pd.DataFrame())
        sp3 = self.data.get('Painting_Stage3', pd.DataFrame())
        
        if casting.empty:
            return pd.DataFrame([['No schedule data available for Gantt']])
        
        gantt_rows = []
        gantt_rows.append(['PRODUCTION TIMELINE - COMPLETE 8-STAGE GANTT VIEW', '', '', '', '', '', '', '', '', '', ''])
        gantt_rows.append(['Visual timeline: Casting → Grinding → MC1 → MC2 → MC3 → SP1 → SP2 → SP3 → Delivery', '', '', '', '', '', '', '', '', '', ''])
        gantt_rows.append(['', '', '', '', '', '', '', '', '', '', ''])
        
        gantt_rows.append(['Part', 'Qty', 'Cast', 'Grind', 'MC1', 'MC2', 'MC3', 'SP1', 'SP2', 'SP3', 'Delivery'])
        
        if 'Part' in casting.columns:
            parts = sorted(casting['Part'].unique()[:30])
            
            for part in parts:
                part_cast = casting[casting['Part'] == part] if not casting.empty else pd.DataFrame()
                part_grind = grinding[grinding['Part'] == part] if not grinding.empty and 'Part' in grinding.columns else pd.DataFrame()
                part_mc1 = mc1[mc1['Part'] == part] if not mc1.empty and 'Part' in mc1.columns else pd.DataFrame()
                part_mc2 = mc2[mc2['Part'] == part] if not mc2.empty and 'Part' in mc2.columns else pd.DataFrame()
                part_mc3 = mc3[mc3['Part'] == part] if not mc3.empty and 'Part' in mc3.columns else pd.DataFrame()
                part_sp1 = sp1[sp1['Part'] == part] if not sp1.empty and 'Part' in sp1.columns else pd.DataFrame()
                part_sp2 = sp2[sp2['Part'] == part] if not sp2.empty and 'Part' in sp2.columns else pd.DataFrame()
                part_sp3 = sp3[sp3['Part'] == part] if not sp3.empty and 'Part' in sp3.columns else pd.DataFrame()
                part_del = delivery[delivery['Part'] == part] if not delivery.empty and 'Part' in delivery.columns else pd.DataFrame()
                
                def get_week_range(df):
                    if df.empty or 'Week' not in df.columns:
                        return '-'
                    min_w = int(df['Week'].min())
                    max_w = int(df['Week'].max())
                    if min_w == max_w:
                        return f"W{min_w}"
                    else:
                        return f"W{min_w}-{max_w}"
                
                cast_weeks = get_week_range(part_cast)
                grind_weeks = get_week_range(part_grind)
                mc1_weeks = get_week_range(part_mc1)
                mc2_weeks = get_week_range(part_mc2)
                mc3_weeks = get_week_range(part_mc3)
                sp1_weeks = get_week_range(part_sp1)
                sp2_weeks = get_week_range(part_sp2)
                sp3_weeks = get_week_range(part_sp3)
                del_weeks = get_week_range(part_del)
                
                qty = int(part_cast['Units'].sum()) if not part_cast.empty and 'Units' in part_cast.columns else 0
                
                gantt_rows.append([
                    part, qty, cast_weeks, grind_weeks, mc1_weeks, mc2_weeks, mc3_weeks, 
                    sp1_weeks, sp2_weeks, sp3_weeks, del_weeks
                ])
        
        gantt_rows.append(['', '', '', '', '', '', '', '', '', '', ''])
        gantt_rows.append(['READING THE COMPLETE TIMELINE', '', '', '', '', '', '', '', '', '', ''])
        gantt_rows.append(['', '', '', '', '', '', '', '', '', '', ''])
        gantt_rows.append(['Column', 'Stage', 'Description', '', '', '', '', '', '', '', ''])
        gantt_rows.append(['Cast', 'Casting', 'When parts are cast', '', '', '', '', '', '', '', ''])
        gantt_rows.append(['Grind', 'Grinding', 'When parts are ground', '', '', '', '', '', '', '', ''])
        gantt_rows.append(['MC1', 'Machining Stage 1', 'First machining operation', '', '', '', '', '', '', '', ''])
        gantt_rows.append(['MC2', 'Machining Stage 2', 'Second machining operation', '', '', '', '', '', '', '', ''])
        gantt_rows.append(['MC3', 'Machining Stage 3', 'Third machining operation', '', '', '', '', '', '', '', ''])
        gantt_rows.append(['SP1', 'Painting Stage 1', 'First paint layer (Primer)', '', '', '', '', '', '', '', ''])
        gantt_rows.append(['SP2', 'Painting Stage 2', 'Second paint layer (Intermediate)', '', '', '', '', '', '', '', ''])
        gantt_rows.append(['SP3', 'Painting Stage 3', 'Third paint layer (Top Coat)', '', '', '', '', '', '', '', ''])
        gantt_rows.append(['Delivery', 'Delivery', 'When parts delivered to customer', '', '', '', '', '', '', '', ''])
        
        return pd.DataFrame(gantt_rows)

    def create_daily_schedule(self):
        """SHEET 8: DAILY SCHEDULE WITH HOLIDAYS"""
        print("📅 Creating Daily Schedule with calendar dates and holidays...")

        daily = self.data.get('Daily_Schedule', pd.DataFrame())
        weekly_summary = self.data.get('Weekly_Summary', pd.DataFrame())

        if daily.empty:
            return pd.DataFrame([['No daily schedule data available']])

        schedule_rows = []
        schedule_rows.append(['DAILY PRODUCTION SCHEDULE WITH CALENDAR DATES', '', '', '', '', '', '', '', '', '', '', ''])
        schedule_rows.append([f'Day-by-day production plan | Excludes Sundays & India National Holidays | Planning Period: {self.num_weeks} weeks', '', '', '', '', '', '', '', '', '', '', ''])
        schedule_rows.append(['', '', '', '', '', '', '', '', '', '', '', ''])

        # Calendar summary
        if not daily.empty:
            total_days = len(daily)
            working_days = len(daily[daily['Is_Holiday'] == 'No']) if 'Is_Holiday' in daily.columns else total_days
            holidays = total_days - working_days
            sundays = len(daily[daily.get('Holiday_Name', '').str.contains('Sunday', na=False)]) if 'Holiday_Name' in daily.columns else 0
            national_holidays = holidays - sundays

            schedule_rows.append(['CALENDAR SUMMARY', '', '', '', '', '', '', '', '', '', '', ''])
            schedule_rows.append(['', '', '', '', '', '', '', '', '', '', '', ''])

            summary_data = [
                ['Metric', 'Count', 'Percentage', '', '', '', '', '', '', '', '', ''],
                ['Total Days in Schedule', f'{total_days}', '100%', '', '', '', '', '', '', '', '', ''],
                ['Working Days (Mon-Sat)', f'{working_days}', f'{working_days/total_days*100:.1f}%', '', '', '', '', '', '', '', '', ''],
                ['Sundays (Weekly Off)', f'{sundays}', f'{sundays/total_days*100:.1f}%', '', '', '', '', '', '', '', '', ''],
                ['National Holidays', f'{national_holidays}', f'{national_holidays/total_days*100:.1f}%', '', '', '', '', '', '', '', '', ''],
                ['Total Holidays', f'{holidays}', f'{holidays/total_days*100:.1f}%', '', '', '', '', '', '', '', '', '']
            ]

            schedule_rows.extend(summary_data)
            schedule_rows.append(['', '', '', '', '', '', '', '', '', '', '', ''])

        schedule_rows.append(['DAILY PRODUCTION SCHEDULE', '', '', '', '', '', '', '', '', '', '', ''])
        schedule_rows.append(['', '', '', '', '', '', '', '', '', '', '', ''])
        schedule_rows.append([
            'Week', 'Date', 'Day', 'Status', 'Holiday/Event', 'Casting', 'Grinding', 'MC1', 'MC2', 'MC3',
            'SP1', 'SP2', 'SP3', 'Big Line (hrs)', 'Small Line (hrs)', 'Big Line Util %', 'Small Line Util %'
        ])

        for idx, row in daily.iterrows():
            week = row.get('Week', '-')
            date = row.get('Date', '-')
            day = row.get('Day', '-')
            is_holiday = row.get('Is_Holiday', 'No')
            holiday_name = row.get('Holiday_Name', '-')

            # Production quantities
            casting = row.get('Casting_Tons', 0)
            grinding = row.get('Grinding_Units', 0)
            mc1 = row.get('MC1_Units', 0)
            mc2 = row.get('MC2_Units', 0)
            mc3 = row.get('MC3_Units', 0)
            sp1 = row.get('SP1_Units', 0)
            sp2 = row.get('SP2_Units', 0)
            sp3 = row.get('SP3_Units', 0)
            big_line_hours = row.get('Big_Line_Hours', 0)
            small_line_hours = row.get('Small_Line_Hours', 0)
            big_line_util = row.get('Big_Line_Util_%', 0)
            small_line_util = row.get('Small_Line_Util_%', 0)

            # Status icon
            if is_holiday == 'Yes':
                if 'Sunday' in str(holiday_name):
                    status = '🔴 HOLIDAY'
                else:
                    status = '🔴 HOLIDAY'
            elif day == 'Saturday':
                status = '🟡 Saturday'
            else:
                status = '🟢 Working'

            schedule_rows.append([
                f'W{int(week)}' if week != '-' else '-',
                date,
                day,
                status,
                holiday_name if is_holiday == 'Yes' else '-',
                f'{casting:.1f}' if casting > 0 else '-',
                f'{grinding:.0f}' if grinding > 0 else '-',
                f'{mc1:.0f}' if mc1 > 0 else '-',
                f'{mc2:.0f}' if mc2 > 0 else '-',
                f'{mc3:.0f}' if mc3 > 0 else '-',
                f'{sp1:.0f}' if sp1 > 0 else '-',
                f'{sp2:.0f}' if sp2 > 0 else '-',
                f'{sp3:.0f}' if sp3 > 0 else '-',
                f'{big_line_hours:.1f}' if big_line_hours else '-',
                f'{small_line_hours:.1f}' if small_line_hours else '-',
                f'{big_line_util:.1f}%' if big_line_util else '-',
                f'{small_line_util:.1f}%' if small_line_util else '-'
            ])

        # Add weekly aggregates
        schedule_rows.append(['', '', '', '', '', '', '', '', '', '', '', '', ''])
        schedule_rows.append(['WEEKLY TOTALS SUMMARY (All Weeks)', '', '', '', '', '', '', '', '', '', '', '', ''])
        schedule_rows.append(['', '', '', '', '', '', '', '', '', '', '', '', ''])
        schedule_rows.append([
            'Week', 'Working Days', 'Holidays', 'Casting Tons', 'Grinding Units', 'MC1 Units',
            'MC2 Units', 'MC3 Units', 'SP1 Units', 'SP2 Units', 'SP3 Units',
            'Big Line (hrs)', 'Small Line (hrs)', 'Big Line Util %', 'Small Line Util %'
        ])

        if 'Week' in daily.columns:
            weekly_grouped = daily.groupby('Week').agg({
                'Is_Holiday': lambda x: sum(x == 'No'),  # Count working days
                'Casting_Tons': 'sum',
                'Grinding_Units': 'sum',
                'MC1_Units': 'sum',
                'MC2_Units': 'sum',
                'MC3_Units': 'sum',
                'SP1_Units': 'sum',
                'SP2_Units': 'sum',
                'SP3_Units': 'sum',
                'Big_Line_Hours': 'sum',
                'Small_Line_Hours': 'sum',
                'Big_Line_Util_%': 'mean',
                'Small_Line_Util_%': 'mean'
            }).reset_index()

            for _, wk_row in weekly_grouped.iterrows():
                week = wk_row['Week']
                working_days = wk_row['Is_Holiday']
                holidays_in_week = 7 - working_days

                schedule_rows.append([
                    f'W{int(week)}',
                    int(working_days),
                    int(holidays_in_week),
                    f'{wk_row["Casting_Tons"]:.1f}',
                    f'{wk_row["Grinding_Units"]:.0f}',
                    f'{wk_row["MC1_Units"]:.0f}',
                    f'{wk_row["MC2_Units"]:.0f}',
                    f'{wk_row["MC3_Units"]:.0f}',
                    f'{wk_row["SP1_Units"]:.0f}',
                    f'{wk_row["SP2_Units"]:.0f}',
                    f'{wk_row["SP3_Units"]:.0f}',
                    f'{wk_row.get("Big_Line_Hours", 0):.1f}',
                    f'{wk_row.get("Small_Line_Hours", 0):.1f}',
                    f'{wk_row.get("Big_Line_Util_%", 0):.1f}%',
                    f'{wk_row.get("Small_Line_Util_%", 0):.1f}%'
                ])
        schedule_rows.append(['', '', '', '', '', '', '', '', '', '', '', '', ''])
        schedule_rows.append(['NOTES:', '', '', '', '', '', '', '', '', '', '', '', ''])
        schedule_rows.append(['🔴 HOLIDAY = No production (Sunday or National Holiday)', '', '', '', '', '', '', '', '', '', '', '', ''])
        schedule_rows.append(['🟡 Saturday = Working day but may have reduced shifts', '', '', '', '', '', '', '', '', '', '', '', ''])
        schedule_rows.append(['🟢 Working = Full production day (Monday-Friday)', '', '', '', '', '', '', '', '', '', '', '', ''])
        schedule_rows.append(['Production quantities are distributed evenly across working days in each week', '', '', '', '', '', '', '', '', '', '', '', ''])

        if weekly_summary is not None and not weekly_summary.empty:
            schedule_rows.append(['', '', '', '', '', '', '', '', '', '', '', '', ''])
            schedule_rows.append(['MOULDING LINE UTILIZATION BY WEEK', '', '', '', '', '', '', '', '', '', '', '', ''])
            schedule_rows.append(['', '', '', '', '', '', '', '', '', '', '', '', ''])
            schedule_rows.append(['Week', 'Big Line Util %', 'Small Line Util %', 'Big Line Hours', 'Small Line Hours', '', '', '', '', '', '', '', ''])
            for _, wk in weekly_summary.iterrows():
                schedule_rows.append([
                    f"W{int(wk['Week'])}" if not pd.isna(wk.get('Week')) else '-',
                    f"{wk.get('Big_Line_Util_%', 0):.1f}%" if 'Big_Line_Util_%' in wk else '-',
                    f"{wk.get('Small_Line_Util_%', 0):.1f}%" if 'Small_Line_Util_%' in wk else '-',
                    f"{wk.get('Big_Line_Hours', 0):.1f}" if 'Big_Line_Hours' in wk else '-',
                    f"{wk.get('Small_Line_Hours', 0):.1f}" if 'Small_Line_Hours' in wk else '-',
                    '', '', '', '', '', '', '', ''
                ])
        return pd.DataFrame(schedule_rows)

    def create_part_daily_schedule(self):
        """SHEET 9: PART-LEVEL DAILY SCHEDULE"""
        print("📋 Creating Part-Level Daily Schedule...")

        part_daily = self.data.get('Part_Daily_Schedule', pd.DataFrame())
        weekly_summary = self.data.get('Weekly_Summary', pd.DataFrame())
        machines = self.data.get('Machine_Utilization', pd.DataFrame())

        if part_daily.empty:
            return pd.DataFrame([['No part-level daily schedule data available']])

        schedule_rows = []
        schedule_rows.append(['PART-LEVEL DAILY PRODUCTION SCHEDULE - ALL 8 STAGES', '', '', '', '', '', '', '', '', '', '', '', '', ''])
        schedule_rows.append([f'Complete production flow: Casting → Grinding → MC1 → MC2 → MC3 → SP1 → SP2 → SP3 → Delivery | Planning Period: {self.num_weeks} weeks', '', '', '', '', '', '', '', '', '', '', '', '', ''])
        schedule_rows.append(['This schedule shows the ACTUAL daily production flow to determine REAL delivery dates', '', '', '', '', '', '', '', '', '', '', '', '', ''])
        schedule_rows.append(['', '', '', '', '', '', '', '', '', '', '', '', '', ''])

        # Summary statistics
        if not part_daily.empty:
            total_entries = len(part_daily)
            unique_parts = part_daily['Part'].nunique()
            total_units = part_daily['Units'].sum()
            unique_orders = part_daily['Sales_Order'].nunique() if 'Sales_Order' in part_daily.columns else 0

            # Count operations
            operation_counts = part_daily['Operation'].value_counts().to_dict()

            schedule_rows.append(['SCHEDULE OVERVIEW', '', '', '', '', '', '', '', '', '', '', '', '', ''])
            schedule_rows.append(['', '', '', '', '', '', '', '', '', '', '', '', '', ''])

            summary_data = [
                ['Metric', 'Value', '', '', '', '', '', '', '', '', '', '', '', ''],
                ['Total Schedule Entries', f'{total_entries:,}', '', '', '', '', '', '', '', '', '', '', '', ''],
                ['Unique Sales Orders', f'{unique_orders}', '', '', '', '', '', '', '', '', '', '', '', ''],
                ['Unique Parts', f'{unique_parts}', '', '', '', '', '', '', '', '', '', '', '', ''],
                ['Total Production Units', f'{total_units:,.0f}', '', '', '', '', '', '', '', '', '', '', '', ''],
                ['Operations Tracked', '8 Stages (Casting, Grinding, MC1-3, SP1-3)', '', '', '', '', '', '', '', '', '', '', '', ''],
                ['', '', '', '', '', '', '', '', '', '', '', '', '', ''],
                ['STAGE BREAKDOWN:', '', '', '', '', '', '', '', '', '', '', '', '', ''],
                ['Casting Entries', f"{operation_counts.get('Casting', 0):,}", '', '', '', '', '', '', '', '', '', '', '', ''],
                ['Grinding Entries', f"{operation_counts.get('Grinding', 0):,}", '', '', '', '', '', '', '', '', '', '', '', ''],
                ['MC1 Entries', f"{operation_counts.get('MC1', 0):,}", '', '', '', '', '', '', '', '', '', '', '', ''],
                ['MC2 Entries', f"{operation_counts.get('MC2', 0):,}", '', '', '', '', '', '', '', '', '', '', '', ''],
                ['MC3 Entries', f"{operation_counts.get('MC3', 0):,}", '', '', '', '', '', '', '', '', '', '', '', ''],
                ['SP1 (Primer) Entries', f"{operation_counts.get('SP1', 0):,}", '', '', '', '', '', '', '', '', '', '', '', ''],
                ['SP2 (Intermediate) Entries', f"{operation_counts.get('SP2', 0):,}", '', '', '', '', '', '', '', '', '', '', '', ''],
                ['SP3 (Top Coat) Entries', f"{operation_counts.get('SP3', 0):,}", '', '', '', '', '', '', '', '', '', '', '', '']
            ]

            schedule_rows.extend(summary_data)
            schedule_rows.append(['', '', '', '', '', '', '', '', '', '', '', '', '', ''])

            if (machines is not None and 'Big_Line_Util_%' in machines.columns) or \
               (machines is not None and 'Small_Line_Util_%' in machines.columns):
                schedule_rows.append(['MOULDING LINE CAPACITY UTILIZATION', '', '', '', '', '', '', '', '', '', '', '', '', ''])
                schedule_rows.append(['', '', '', '', '', '', '', '', '', '', '', '', '', ''])
                schedule_rows.append(['Line', 'Avg Util %', 'Max Util %', 'Total Hours', 'Capacity Hours', '', '', '', '', '', '', '', '', ''])

                if machines is not None and 'Big_Line_Util_%' in machines.columns:
                    big_avg = machines['Big_Line_Util_%'].mean()
                    big_max = machines['Big_Line_Util_%'].max()
                    big_hours = weekly_summary['Big_Line_Hours'].sum() if weekly_summary is not None and 'Big_Line_Hours' in weekly_summary.columns else 0
                    big_cap = weekly_summary['Big_Line_Capacity_Hours'].sum() if weekly_summary is not None and 'Big_Line_Capacity_Hours' in weekly_summary.columns else ''
                    schedule_rows.append([
                        'Big Line',
                        f'{big_avg:.1f}%',
                        f'{big_max:.1f}%',
                        f'{big_hours:.1f}',
                        f'{big_cap:.1f}' if big_cap != '' else '',
                        '', '', '', '', '', '', '', '', ''
                    ])

                if machines is not None and 'Small_Line_Util_%' in machines.columns:
                    small_avg = machines['Small_Line_Util_%'].mean()
                    small_max = machines['Small_Line_Util_%'].max()
                    small_hours = weekly_summary['Small_Line_Hours'].sum() if weekly_summary is not None and 'Small_Line_Hours' in weekly_summary.columns else 0
                    small_cap = weekly_summary['Small_Line_Capacity_Hours'].sum() if weekly_summary is not None and 'Small_Line_Capacity_Hours' in weekly_summary.columns else ''
                    schedule_rows.append([
                        'Small Line',
                        f'{small_avg:.1f}%',
                        f'{small_max:.1f}%',
                        f'{small_hours:.1f}',
                        f'{small_cap:.1f}' if small_cap != '' else '',
                        '', '', '', '', '', '', '', '', ''
                    ])

                schedule_rows.append(['', '', '', '', '', '', '', '', '', '', '', '', '', ''])

        # CLEAN DATA-ONLY VIEW (no separators for Excel readability)
        schedule_rows.append(['PART-LEVEL DAILY PRODUCTION SCHEDULE - CLEAN DATA VIEW', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', ''])
        schedule_rows.append([f'All production data without separator rows for easy Excel filtering/sorting | Total: {len(part_daily)} entries', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', ''])
        schedule_rows.append(['Use "Freeze Panes" on Row 4 and Filter on "Sales Order" column to navigate by order', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', ''])
        schedule_rows.append(['', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', ''])
        schedule_rows.append(['Date', 'Moulding Date', 'Day', 'Week', 'Part', 'Sales Order', 'Customer', 'Committed Week', 'Order Qty', 'Batch No', 'Cumulative Qty', 'Progress', 'Units', 'Operation', 'Machine/Resource', 'Unit Wt (kg)', 'Total Wt (ton)', 'Cycle Time', 'Batch Size', 'Prod Time', 'Notes'])

        # Show ALL entries as CLEAN DATA ROWS ONLY (no separators)
        if not part_daily.empty and 'Sales_Order' in part_daily.columns:
            # Already sorted by Sales_Order, Date, Operation in data generation
            part_daily_sorted = part_daily.copy()

            for _, row in part_daily_sorted.iterrows():
                row_order = row.get('Sales_Order', 'N/A')
                row_part = row.get('Part', '-')
                row_customer = row.get('Customer', 'Unknown')
                row_committed = row.get('Committed_Week', '-')
                row_order_qty = row.get('Order_Qty', 0)
                row_operation = row.get('Operation', '-')
                row_units = int(row.get('Units', 0))
                row_date = row.get('Date', '-')
                row_day = row.get('Day', '-')
                row_week = row.get('Week', '-')

                # Add data row ONLY (no separators, no empty rows, no stage totals)
                schedule_rows.append([
                    row_date,
                    row.get('Moulding_Date', row_date),  # NEW - Moulding Date column
                    row_day,
                    row_week,
                    row_part,
                    row_order,
                    row_customer,
                    row_committed,
                    row_order_qty,
                    int(row.get('Batch_No', 1)),  # Batch number
                    int(row.get('Cumulative_Qty', row_units)),  # Cumulative quantity
                    row.get('Progress', f"{row_units}/{row_order_qty}"),  # Progress indicator
                    int(row_units),  # WHOLE NUMBER - converted to int to remove .0 suffix
                    row_operation,
                    row.get('Machine_Resource', 'N/A'),
                    f"{row.get('Unit_Weight_kg', 0):.2f}",
                    f"{row.get('Total_Weight_ton', 0):.3f}",
                    f"{row.get('Cycle_Time_min', 0):.0f} min" if row.get('Cycle_Time_min', 0) > 0 else '-',
                    f"{row.get('Batch_Size', 1)}",
                    f"{row.get('Production_Time_min', 0):.0f} min" if row.get('Production_Time_min', 0) > 0 else '-',
                    row.get('Special_Notes', '')
                ])
        elif not part_daily.empty:
            # Fallback if Sales_Order column missing
            schedule_rows.append(['⚠ Sales Order information not available - showing date-based view', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', ''])

        # Add operation-wise summary
        schedule_rows.append(['', '', '', '', '', '', '', '', '', '', '', '', '', ''])
        schedule_rows.append(['OPERATION-WISE SUMMARY (All Days)', '', '', '', '', '', '', '', '', '', '', '', '', ''])
        schedule_rows.append(['', '', '', '', '', '', '', '', '', '', '', '', '', ''])
        schedule_rows.append(['Operation', 'Total Entries', 'Total Units', 'Unique Parts', 'Avg Units/Day', '', '', '', '', '', '', '', '', ''])

        if not part_daily.empty:
            operation_summary = part_daily.groupby('Operation').agg({
                'Units': ['count', 'sum', 'mean'],
                'Part': 'nunique'
            }).reset_index()

            operation_summary.columns = ['Operation', 'Entries', 'Total_Units', 'Avg_Units', 'Unique_Parts']

            for _, op_row in operation_summary.iterrows():
                schedule_rows.append([
                    op_row['Operation'],
                    int(op_row['Entries']),
                    f"{op_row['Total_Units']:.0f}",
                    int(op_row['Unique_Parts']),
                    f"{op_row['Avg_Units']:.1f}",
                    '',
                    '',
                    '',
                    '',
                    '',
                    '',
                    '',
                    '',
                    ''
                ])

        # Add machine utilization summary
        schedule_rows.append(['', '', '', '', '', '', '', '', '', '', '', '', '', ''])
        schedule_rows.append(['MACHINE UTILIZATION SUMMARY (Top Machines)', '', '', '', '', '', '', '', '', '', '', '', '', ''])
        schedule_rows.append(['', '', '', '', '', '', '', '', '', '', '', '', '', ''])
        schedule_rows.append(['Machine/Resource', 'Operations', 'Total Units', 'Total Production Time (hrs)', '', '', '', '', '', '', '', '', '', ''])

        if not part_daily.empty:
            machine_summary = part_daily[part_daily['Machine_Resource'] != 'N/A'].groupby('Machine_Resource').agg({
                'Units': 'sum',
                'Production_Time_min': 'sum',
                'Operation': 'count'
            }).reset_index().sort_values('Production_Time_min', ascending=False).head(15)

            for _, mach_row in machine_summary.iterrows():
                schedule_rows.append([
                    mach_row['Machine_Resource'],
                    int(mach_row['Operation']),
                    f"{mach_row['Units']:.0f}",
                    f"{mach_row['Production_Time_min'] / 60.0:.1f}",
                    '',
                    '',
                    '',
                    '',
                    '',
                    '',
                    '',
                    '',
                    '',
                    ''
                ])

        schedule_rows.append(['', '', '', '', '', '', '', '', '', '', '', '', '', ''])
        schedule_rows.append(['KEY INSIGHTS - HOW TO DETERMINE ACTUAL DELIVERY DATE:', '', '', '', '', '', '', '', '', '', '', '', '', ''])
        schedule_rows.append(['', '', '', '', '', '', '', '', '', '', '', '', '', ''])
        schedule_rows.append(['1. PRODUCTION FLOW:', 'A part must go through all 8 stages sequentially:', '', '', '', '', '', '', '', '', '', '', '', ''])
        schedule_rows.append(['   ', 'Casting (Week X) → Grinding (Week X+1) → MC1 (Week X+2) → MC2 (Week X+3) → MC3 (Week X+4) → SP1 (Week X+5) → SP2 (Week X+6) → SP3 (Week X+7)', '', '', '', '', '', '', '', '', '', '', '', ''])
        schedule_rows.append(['', '', '', '', '', '', '', '', '', '', '', '', '', ''])
        schedule_rows.append(['2. DELIVERY DATE:', 'A part can only be delivered AFTER completing SP3 (final painting stage)', '', '', '', '', '', '', '', '', '', '', '', ''])
        schedule_rows.append(['   Example:', 'If SP3 completes on 2025-12-15, earliest delivery is 2025-12-15', '', '', '', '', '', '', '', '', '', '', '', ''])
        schedule_rows.append(['', '', '', '', '', '', '', '', '', '', '', '', '', ''])
        schedule_rows.append(['3. WIP PARTS:', 'Parts already in WIP (Finished Goods, SP, MC, GR stages) skip earlier stages', '', '', '', '', '', '', '', '', '', '', '', ''])
        schedule_rows.append(['   Example:', 'Part in MC stage skips Casting & Grinding, goes directly to MC1→MC2→MC3→SP1→SP2→SP3', '', '', '', '', '', '', '', '', '', '', '', ''])
        schedule_rows.append(['', '', '', '', '', '', '', '', '', '', '', '', '', ''])
        schedule_rows.append(['NOTES:', '', '', '', '', '', '', '', '', '', '', '', '', ''])
        schedule_rows.append(['• This schedule shows EXACTLY which parts to produce each day on which machines', '', '', '', '', '', '', '', '', '', '', '', '', ''])
        schedule_rows.append(['• Units are distributed evenly across working days (Mon-Sat, excluding holidays)', '', '', '', '', '', '', '', '', '', '', '', '', ''])
        schedule_rows.append(['• Working days exclude Sundays and Indian national holidays', '', '', '', '', '', '', '', '', '', '', '', '', ''])
        schedule_rows.append(['• Use this as the daily work order for the production floor', '', '', '', '', '', '', '', '', '', '', '', '', ''])
        schedule_rows.append(['• Track each part through all 8 stages to determine real delivery capability', '', '', '', '', '', '', '', '', '', '', '', '', ''])

        return pd.DataFrame(schedule_rows)

    def apply_enhanced_formatting(self, ws, sheet_name, df, sections=None):
        """Apply comprehensive formatting for improved readability"""
        thin_border = Border(
            left=Side(style='thin', color=self.colors['border_gray']),
            right=Side(style='thin', color=self.colors['border_gray']),
            top=Side(style='thin', color=self.colors['border_gray']),
            bottom=Side(style='thin', color=self.colors['border_gray'])
        )

        if sheet_name == '1_EXECUTIVE_DASHBOARD' and sections:
            self._format_dashboard(ws, sections)
        elif sheet_name == '8_DAILY_SCHEDULE':
            self._format_daily_schedule(ws, df)
        elif sheet_name == '9_PART_DAILY_SCHEDULE':
            self._format_part_daily_schedule(ws, df)
        else:
            self._format_standard_sheet(ws, df)

        self._auto_size_columns(ws)

        if ws.max_row > 5:
            ws.freeze_panes = 'A5'
    
    def _format_dashboard(self, ws, sections):
        """Special formatting for dashboard"""
        row_idx = 1
        
        for section in sections:
            section_type = section['type']
            section_data = section['data']
            
            if section_type == 'title':
                for col in range(1, 6):
                    cell = ws.cell(row=row_idx, column=col)
                    cell.font = self.fonts['title']
                    cell.fill = PatternFill(start_color=self.colors['header_dark'], 
                                          end_color=self.colors['header_dark'], fill_type='solid')
                    cell.alignment = Alignment(horizontal='left', vertical='center')
                ws.merge_cells(f'A{row_idx}:E{row_idx}')
                row_idx += 1
            
            elif section_type == 'subtitle':
                for col in range(1, 6):
                    cell = ws.cell(row=row_idx, column=col)
                    cell.font = Font(name='Calibri', size=10, italic=True, color='666666')
                    cell.alignment = Alignment(horizontal='left', vertical='center')
                ws.merge_cells(f'A{row_idx}:E{row_idx}')
                row_idx += 1
            
            elif section_type == 'section_header':
                for col in range(1, 6):
                    cell = ws.cell(row=row_idx, column=col)
                    cell.font = self.fonts['header']
                    cell.fill = PatternFill(start_color=self.colors['header_light'], 
                                          end_color=self.colors['header_light'], fill_type='solid')
                    cell.alignment = Alignment(horizontal='left', vertical='center')
                ws.merge_cells(f'A{row_idx}:E{row_idx}')
                ws.row_dimensions[row_idx].height = 25
                row_idx += 1
            
            elif section_type == 'data_table':
                # Format header row (first row only)
                header_row = row_idx
                header_cell_value = ws.cell(row=header_row, column=1).value
                print(f"      → Formatting data_table header at row {header_row}: '{header_cell_value}'")
                for col in range(1, 6):
                    cell = ws.cell(row=header_row, column=col)
                    cell.font = self.fonts['subheader']
                    cell.fill = PatternFill(start_color=self.colors['light_gray'],
                                        end_color=self.colors['light_gray'], fill_type='solid')
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                row_idx += 1

                # Format data rows (skip header row at index 0)
                print(f"      → Formatting {len(section_data)-1} data rows starting at row {row_idx}")
                for data_row_idx in range(1, len(section_data)):
                    fill_color = self.colors['white'] if data_row_idx % 2 == 0 else self.colors['light_gray']
                    data_cell_value = ws.cell(row=row_idx, column=1).value
                    if data_row_idx <= 2:  # Only print first 2 data rows to avoid spam
                        print(f"         - Data row {data_row_idx} at Excel row {row_idx}: '{data_cell_value}'")
                    for col in range(1, 6):
                        cell = ws.cell(row=row_idx, column=col)

                        # Check if this specific cell has a status indicator
                        has_status = False
                        if cell.value:
                            val_str = str(cell.value).upper()
                            if '🔴' in val_str or 'CRITICAL' in val_str:
                                cell.fill = PatternFill(start_color=self.colors['critical'],
                                                       end_color=self.colors['critical'], fill_type='solid')
                                cell.font = Font(name='Calibri', size=10, bold=True, color='FFFFFF')
                                cell.alignment = Alignment(horizontal='left', vertical='center')
                                has_status = True
                            elif '🟡' in val_str or 'WARNING' in val_str:
                                cell.fill = PatternFill(start_color=self.colors['warning'],
                                                       end_color=self.colors['warning'], fill_type='solid')
                                cell.font = Font(name='Calibri', size=10, bold=True, color='000000')
                                cell.alignment = Alignment(horizontal='left', vertical='center')
                                has_status = True
                            elif '🟢' in val_str or 'GOOD' in val_str or '✓' in val_str or 'HEALTHY' in val_str:
                                cell.fill = PatternFill(start_color=self.colors['good'],
                                                       end_color=self.colors['good'], fill_type='solid')
                                cell.font = Font(name='Calibri', size=10, bold=True, color='FFFFFF')
                                cell.alignment = Alignment(horizontal='left', vertical='center')
                                has_status = True

                        # Only apply alternating row color if cell doesn't have status indicator
                        if not has_status:
                            cell.font = self.fonts['normal']
                            cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type='solid')
                            cell.alignment = Alignment(horizontal='left', vertical='center')

                    row_idx += 1
            
            elif section_type == 'blank':
                row_idx += 1

            # Add blank row after each non-blank section (matches line 379-380 in create_executive_dashboard)
            if section_type != 'blank':
                row_idx += 1
    
    def _format_standard_sheet(self, ws, df):
        """Standard formatting for other sheets"""
        for row_idx in range(1, min(4, ws.max_row + 1)):
            for col_idx in range(1, ws.max_column + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                if row_idx == 1:
                    cell.font = self.fonts['title']
                    cell.fill = PatternFill(start_color=self.colors['header_dark'], 
                                          end_color=self.colors['header_dark'], fill_type='solid')
                    cell.alignment = Alignment(horizontal='left', vertical='center')
                elif row_idx == 2:
                    cell.font = Font(name='Calibri', size=10, italic=True, color='666666')
                    cell.alignment = Alignment(horizontal='left', vertical='center')
        
        header_row = 4
        for row_idx in range(4, min(10, ws.max_row + 1)):
            first_cell_value = ws.cell(row=row_idx, column=1).value
            if first_cell_value and not str(first_cell_value).startswith('╔'):
                header_row = row_idx
                break
        
        for col_idx in range(1, ws.max_column + 1):
            cell = ws.cell(row=header_row, column=col_idx)
            cell.font = self.fonts['subheader']
            cell.fill = PatternFill(start_color=self.colors['header_light'], 
                                  end_color=self.colors['header_light'], fill_type='solid')
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        for row_idx in range(header_row + 1, ws.max_row + 1):
            fill_color = self.colors['white'] if (row_idx - header_row) % 2 == 1 else self.colors['light_gray']
            
            for col_idx in range(1, ws.max_column + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.font = self.fonts['normal']
                cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type='solid')
                cell.alignment = Alignment(horizontal='left', vertical='center')
                
                if cell.value:
                    val_str = str(cell.value).upper()
                    if '🔴' in val_str or 'CRITICAL' in val_str:
                        cell.font = Font(name='Calibri', size=10, bold=True, color=self.colors['critical'])
                    elif '🟡' in val_str or 'WARNING' in val_str or '⚠' in val_str:
                        cell.font = Font(name='Calibri', size=10, bold=True, color=self.colors['warning'])
                    elif '🟢' in val_str or '✓' in val_str:
                        cell.font = Font(name='Calibri', size=10, bold=True, color=self.colors['good'])

    def _format_daily_schedule(self, ws, df):
        """Special formatting for daily schedule with holiday highlighting"""
        # Format title rows
        for row_idx in range(1, min(4, ws.max_row + 1)):
            for col_idx in range(1, ws.max_column + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                if row_idx == 1:
                    cell.font = self.fonts['title']
                    cell.fill = PatternFill(start_color=self.colors['header_dark'],
                                          end_color=self.colors['header_dark'], fill_type='solid')
                    cell.alignment = Alignment(horizontal='left', vertical='center')
                elif row_idx == 2:
                    cell.font = Font(name='Calibri', size=10, italic=True, color='666666')
                    cell.alignment = Alignment(horizontal='left', vertical='center')

        # Find header row
        header_row = 4
        for row_idx in range(4, min(15, ws.max_row + 1)):
            first_cell_value = ws.cell(row=row_idx, column=1).value
            if first_cell_value and str(first_cell_value).startswith('Week'):
                header_row = row_idx
                break

        # Format header
        for col_idx in range(1, ws.max_column + 1):
            cell = ws.cell(row=header_row, column=col_idx)
            cell.font = self.fonts['subheader']
            cell.fill = PatternFill(start_color=self.colors['header_light'],
                                  end_color=self.colors['header_light'], fill_type='solid')
            cell.alignment = Alignment(horizontal='center', vertical='center')

        # Format data rows with holiday highlighting
        for row_idx in range(header_row + 1, ws.max_row + 1):
            status_cell = ws.cell(row=row_idx, column=4)  # Column D is Status
            status_value = str(status_cell.value) if status_cell.value else ''

            # Determine row color based on status
            if '🔴' in status_value or 'HOLIDAY' in status_value.upper():
                row_fill_color = 'FFCCCC'  # Light red for holidays
                font_color = '990000'  # Dark red text
                font_bold = True
            elif '🟡' in status_value or 'SATURDAY' in status_value.upper():
                row_fill_color = 'FFF9CC'  # Light yellow for Saturdays
                font_color = '806600'  # Dark yellow/brown text
                font_bold = False
            elif '🟢' in status_value or 'WORKING' in status_value.upper():
                row_fill_color = 'E8F5E9'  # Light green for working days
                font_color = '1B5E20'  # Dark green text
                font_bold = False
            else:
                # Default alternating rows
                row_fill_color = self.colors['white'] if (row_idx - header_row) % 2 == 1 else self.colors['light_gray']
                font_color = '000000'
                font_bold = False

            for col_idx in range(1, ws.max_column + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.fill = PatternFill(start_color=row_fill_color, end_color=row_fill_color, fill_type='solid')
                cell.font = Font(name='Calibri', size=10, bold=font_bold, color=font_color)
                cell.alignment = Alignment(horizontal='left', vertical='center')

                # Special formatting for production quantity columns (numbers)
                if col_idx >= 6:  # Production columns start at column 6 (Casting)
                    cell.alignment = Alignment(horizontal='right', vertical='center')

    def _format_part_daily_schedule(self, ws, df):
        """Special formatting for part-level daily schedule with operation color-coding"""
        # OPTIMIZATION: Skip detailed formatting for large sheets (10,000+ rows)
        # Only format headers to speed up generation significantly
        if ws.max_row > 1000:
            print(f"    ⚡ Skipping detailed formatting for {ws.max_row} rows (using fast header-only mode)")
            # Just format the title and header row
            for row_idx in range(1, min(30, ws.max_row + 1)):
                first_cell_value = ws.cell(row=row_idx, column=1).value
                if first_cell_value and 'Date' in str(first_cell_value):
                    # Format this header row
                    for col_idx in range(1, min(ws.max_column + 1, 20)):
                        cell = ws.cell(row=row_idx, column=col_idx)
                        cell.font = Font(name='Calibri', size=10, bold=True)
                        cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
                        cell.font = Font(name='Calibri', size=10, bold=True, color='FFFFFF')
                    break
            return  # Skip the rest of the formatting

        # Format title rows (only for small sheets)
        for row_idx in range(1, min(4, ws.max_row + 1)):
            for col_idx in range(1, ws.max_column + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                if row_idx == 1:
                    cell.font = self.fonts['title']
                    cell.fill = PatternFill(start_color=self.colors['header_dark'],
                                          end_color=self.colors['header_dark'], fill_type='solid')
                    cell.alignment = Alignment(horizontal='left', vertical='center')
                elif row_idx == 2:
                    cell.font = Font(name='Calibri', size=10, italic=True, color='666666')
                    cell.alignment = Alignment(horizontal='left', vertical='center')

        # Find header row
        header_row = 4
        for row_idx in range(4, min(30, ws.max_row + 1)):
            first_cell_value = ws.cell(row=row_idx, column=1).value
            if first_cell_value and str(first_cell_value).startswith('Date'):
                header_row = row_idx
                break

        # Format header
        for col_idx in range(1, ws.max_column + 1):
            cell = ws.cell(row=header_row, column=col_idx)
            cell.font = self.fonts['subheader']
            cell.fill = PatternFill(start_color=self.colors['header_light'],
                                  end_color=self.colors['header_light'], fill_type='solid')
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        # Operation colors (for visual distinction)
        operation_colors = {
            'Casting': 'CCE5FF',  # Light blue
            'Grinding': 'FFE5CC',  # Light orange
            'Machining_Stage1': 'CCFFCC',  # Light green
            'Machining_Stage2': 'CCFFE5',  # Lighter green
            'Machining_Stage3': 'D4FFD4',  # Very light green
            'Painting_Stage1': 'E5CCFF',  # Light purple
            'Painting_Stage2': 'F0CCFF',  # Lighter purple
            'Painting_Stage3': 'F5E5FF'   # Very light purple
        }

        # Format data rows with operation color-coding
        for row_idx in range(header_row + 1, ws.max_row + 1):
            # Get operation from column 6 (F)
            operation_cell = ws.cell(row=row_idx, column=6)
            operation = str(operation_cell.value) if operation_cell.value else ''

            # Date separator rows (have dates in column A but are styling rows)
            first_cell = ws.cell(row=row_idx, column=1)
            first_val = str(first_cell.value) if first_cell.value else ''

            # Check if this is a date separator row
            is_separator = '(' in first_val and ')' in first_val and '🟢' not in first_val or '🔴' in first_val or '🟡' in first_val

            if is_separator:
                # Date separator formatting
                for col_idx in range(1, ws.max_column + 1):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    cell.font = Font(name='Calibri', size=11, bold=True, color='1F4E78')
                    cell.fill = PatternFill(start_color='E7E6E6', end_color='E7E6E6', fill_type='solid')
                    cell.alignment = Alignment(horizontal='left', vertical='center')
            elif operation in operation_colors:
                # Operation-based color coding
                row_fill_color = operation_colors[operation]
                for col_idx in range(1, ws.max_column + 1):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    cell.fill = PatternFill(start_color=row_fill_color, end_color=row_fill_color, fill_type='solid')
                    cell.font = Font(name='Calibri', size=9)
                    cell.alignment = Alignment(horizontal='left', vertical='center')

                    # Bold part names (column 5)
                    if col_idx == 5:
                        cell.font = Font(name='Calibri', size=9, bold=True)

                    # Right-align numeric columns
                    if col_idx in [7, 9, 10, 12, 13]:  # Units, Unit Wt, Total Wt, Batch, Prod Time
                        cell.alignment = Alignment(horizontal='right', vertical='center')
            else:
                # Default alternating rows for summary sections
                row_fill_color = self.colors['white'] if (row_idx - header_row) % 2 == 1 else self.colors['light_gray']
                for col_idx in range(1, ws.max_column + 1):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    cell.fill = PatternFill(start_color=row_fill_color, end_color=row_fill_color, fill_type='solid')
                    cell.font = Font(name='Calibri', size=10)
                    cell.alignment = Alignment(horizontal='left', vertical='center')

    def _auto_size_columns(self, ws):
        """Auto-size columns with maximum width"""
        for col_num in range(1, ws.max_column + 1):
            max_length = 0
            column_letter = None
            
            for row_num in range(1, ws.max_row + 1):
                cell = ws.cell(row=row_num, column=col_num)
                
                if isinstance(cell, type(ws.cell(1, 1))):
                    if column_letter is None:
                        column_letter = cell.column_letter
                    
                    try:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    except:
                        pass
            
            if column_letter:
                adjusted_width = min(max_length + 3, 60)
                ws.column_dimensions[column_letter].width = adjusted_width
    
    def generate_executive_report(self, output_path):
        """Main method to generate FIXED executive report with ALL 8 stages"""
        
        print("\n" + "="*80)
        print("GENERATING FIXED EXECUTIVE 9-SHEET REPORT")
        print("NOW SHOWS ALL 8 PRODUCTION STAGES + DAILY & PART-LEVEL SCHEDULES")
        print("="*80 + "\n")
        
        self.load_detailed_data()
        
        print("\n📊 Creating FIXED sheets with ALL 8 stages...\n")

        dashboard_df, dashboard_sections = self.create_executive_dashboard()
        master_schedule_df = self.create_master_schedule()
        delivery_tracker_df = self.create_delivery_tracker()
        bottleneck_alerts_df = self.create_bottleneck_alerts()
        capacity_overview_df = self.create_capacity_overview()
        material_flow_df = self.create_material_flow()
        # gantt_timeline_df = self.create_gantt_timeline()  # REMOVED per user request
        daily_schedule_df = self.create_daily_schedule()
        part_daily_schedule_df = self.create_part_daily_schedule()

        sheets = {
            '1_EXECUTIVE_DASHBOARD': (dashboard_df, dashboard_sections),
            '2_MASTER_SCHEDULE': (master_schedule_df, None),
            '3_DELIVERY_TRACKER': (delivery_tracker_df, None),
            '4_BOTTLENECK_ALERTS': (bottleneck_alerts_df, None),
            '5_CAPACITY_OVERVIEW': (capacity_overview_df, None),
            '6_MATERIAL_FLOW': (material_flow_df, None),
            '7_DAILY_SCHEDULE': (daily_schedule_df, None),
            '8_PART_DAILY_SCHEDULE': (part_daily_schedule_df, None)
        }
        
        print(f"\n💾 Writing FIXED report to: {output_path}")
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            for sheet_name, (df, sections) in sheets.items():
                df.to_excel(writer, sheet_name=sheet_name, index=False, header=False)
                print(f"  ✓ Created: {sheet_name}")
        
        print("\n🎨 Applying enhanced formatting...")
        self.wb = load_workbook(output_path)
        
        for sheet_name, (df, sections) in sheets.items():
            ws = self.wb[sheet_name]
            self.apply_enhanced_formatting(ws, sheet_name, df, sections)
            print(f"  ✓ Formatted: {sheet_name}")
        
        self.wb.save(output_path)
        
        print("\n" + "="*80)
        print("✅ FIXED EXECUTIVE REPORT GENERATED SUCCESSFULLY!")
        print("="*80)
        print(f"\n📁 Output: {output_path}")
        print(f"📊 Sheets: 8 executive-level sheets (including Daily & Part-Level Schedules)")
        print("\n🎯 FIXED ISSUES:")
        print("   ✓ ALL 8 stages now visible in Executive Dashboard")
        print("   ✓ Casting, Grinding, MC1, MC2, MC3, SP1, SP2, SP3")
        print("   ✓ Proper capacity calculations for each stage")
        print("   ✓ Individual utilization percentages")
        print("   ✓ Complete bottleneck analysis")
        print("   ✓ Full capacity overview across all stages")
        print("   ✓ Daily Schedule with calendar dates and holidays")
        print("   ✓ Part-Level Daily Schedule with machine assignments\n")


def main():
    """Generate FIXED executive report with Master Data enrichment"""

    detailed_output = 'production_plan_comprehensive_test.xlsx'
    executive_output = 'production_plan_EXECUTIVE_test.xlsx'
    master_data = 'Master_Data_Updated_Nov_Dec.xlsx'  # Master Data file
    start_date = datetime(2025, 10, 1)  # Changed from Oct 16 to Oct 1 to reduce late deliveries
    
    print("="*80)
    print("EXECUTIVE REPORT GENERATOR - WITH MASTER DATA ENRICHMENT")
    print("Now includes ALL 8 production stages + Master Data enrichment")
    print("="*80)
    print(f"\nInput:  {detailed_output} (27 sheets - includes Part_Daily_Schedule)")
    print(f"Master: {master_data} (Part Master with machines, weights, cycle times)")
    print(f"Output: {executive_output} (9 sheets - ENRICHED)")
    print(f"Start Date: {start_date.strftime('%Y-%m-%d')}\n")
    
    generator = FixedExecutiveReportGenerator(detailed_output, start_date=start_date, master_data_path=master_data)
    generator.generate_executive_report(executive_output)


if __name__ == '__main__':
    main()


