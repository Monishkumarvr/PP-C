import pandas as pd
import openpyxl
from openpyxl import load_workbook

# Read Part Master
df = pd.read_excel('Master_Data_Updated_Nov_Dec.xlsx', sheet_name='Part Master')

print('='*80)
print('FIXING RESOURCE NAME TYPOS IN PART MASTER')
print('='*80)

# Count parts with BVCAD before fix
bvcad_count_before = (
    (df['Machining resource code 1'].str.contains('BVCAD', na=False)).sum() +
    (df['Machining resource code 2'].str.contains('BVCAD', na=False)).sum() +
    (df['Machining resource code 3'].str.contains('BVCAD', na=False)).sum()
)

print(f'\nParts using BVCAD before fix: {bvcad_count_before}')

# Fix 1: Replace BVCAD with KVCAD in machining resource codes
df['Machining resource code 1'] = df['Machining resource code 1'].str.replace('BVCAD', 'KVCAD', regex=False)
df['Machining resource code 2'] = df['Machining resource code 2'].str.replace('BVCAD', 'KVCAD', regex=False)
df['Machining resource code 3'] = df['Machining resource code 3'].str.replace('BVCAD', 'KVCAD', regex=False)

# Count after fix
bvcad_count_after = (
    (df['Machining resource code 1'].str.contains('BVCAD', na=False)).sum() +
    (df['Machining resource code 2'].str.contains('BVCAD', na=False)).sum() +
    (df['Machining resource code 3'].str.contains('BVCAD', na=False)).sum()
)

print(f'Parts using BVCAD after fix: {bvcad_count_after}')
print(f'\n✓ Fixed {bvcad_count_before - bvcad_count_after} resource name typos')

# Write back to Excel
wb = load_workbook('Master_Data_Updated_Nov_Dec.xlsx')
ws = wb['Part Master']

# Clear existing data (keep header)
for row in range(2, ws.max_row + 1):
    for col in range(1, ws.max_column + 1):
        ws.cell(row, col).value = None

# Write updated data
for col_idx, col_name in enumerate(df.columns, 1):
    ws.cell(1, col_idx, col_name)

for row_idx, row in enumerate(df.itertuples(index=False), 2):
    for col_idx, value in enumerate(row, 1):
        ws.cell(row_idx, col_idx, value)

wb.save('Master_Data_Updated_Nov_Dec.xlsx')

print('\n✓ Updated Part Master sheet in Master_Data_Updated_Nov_Dec.xlsx')

# Show affected parts
print('\n' + '='*80)
print('PARTS THAT WERE FIXED:')
print('='*80)
fixed_parts = df[
    (df['Machining resource code 1'].str.contains('KVCAD30MC001|KVCAD40MC001', na=False)) |
    (df['Machining resource code 2'].str.contains('KVCAD30MC001|KVCAD40MC001', na=False))
]

for _, p in fixed_parts.iterrows():
    mc1 = p.get('Machining resource code 1', '')
    mc2 = p.get('Machining resource code 2', '')
    print(f'{p["FG Code"]:15} | MC1={mc1:20} MC2={mc2}')

print(f'\n✓ Total parts fixed: {len(fixed_parts)}')
